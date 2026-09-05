from datetime import datetime, date, time, timedelta
from decimal import Decimal
from io import BytesIO

from extensions import db
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Index
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.orm import validates


def _summary(item, fields):
    if not item:
        return None
    return {field: getattr(item, field) for field in fields}


class TimestampMixin(object):

    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        data = {}
        for column in self.__table__.columns:
            value = getattr(self, column.name)
            if isinstance(value, datetime):
                data[column.name] = value.isoformat()
            elif isinstance(value, date):
                data[column.name] = value.isoformat()
            elif isinstance(value, time):
                data[column.name] = value.isoformat()
            elif isinstance(value, Decimal):
                data[column.name] = float(value)
            else:
                data[column.name] = value
        return data

    @classmethod
    def get_next_id(cls, active_only=False):
        query = db.session.query(cls.id)
        if active_only and hasattr(cls, "is_active"):
            query = query.filter(cls.is_active == True)

        ids = [row.id for row in query.order_by(cls.id).all()]
        next_id = 1
        for current_id in ids:
            if current_id != next_id:
                break
            next_id += 1
        return next_id


class Role(TimestampMixin, db.Model):

    __tablename__ = "roles"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True, index=True)
    # Master category this role belongs to (Admin / HR / Employee / Finance) —
    # drives the master/sub-master grouping on the Roles screen. Free-form
    # roles created under a sub-master category inherit it automatically.
    category = db.Column(db.String(20), nullable=False, default="HR", server_default="HR")
    is_active = db.Column(db.Boolean, default=True)
    actions = db.Column(db.Text, nullable=True)
    users = db.relationship("BaseUser", back_populates="role_obj", cascade="all, delete-orphan")
    permission_links = db.relationship("RolePermission", back_populates="role", cascade="all, delete-orphan")

    def __repr__(self):
        return f"<Role {self.name}>"

    def to_dict(self):
        data = super().to_dict()
        data["permissions"] = [
            _summary(link.permission, ["id", "module", "action"])
            for link in self.permission_links
        ]
        return data


class Permission(TimestampMixin, db.Model):
    """One row per (module, action) pair, e.g. ("Employee", "add").
    Seeded for the six modules x four actions; see migrations for the seed."""

    __tablename__ = "permissions"

    id = db.Column(db.Integer, primary_key=True)
    module = db.Column(db.String(50), nullable=False)
    action = db.Column(db.String(20), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    role_links = db.relationship("RolePermission", back_populates="permission", cascade="all, delete-orphan")

    __table_args__ = (db.UniqueConstraint("module", "action", name="uq_permission_module_action"),)

    def to_dict(self):
        return super().to_dict()


class RolePermission(TimestampMixin, db.Model):
    """Join table granting a Permission to a Role — replaces the free-form
    Role.actions text field with a real matrix for the Roles & Permissions
    screen. Role.actions is left in place for backward compatibility."""

    __tablename__ = "role_permissions"

    id = db.Column(db.Integer, primary_key=True)
    role_id = db.Column(db.Integer, db.ForeignKey("roles.id", ondelete="CASCADE"), nullable=False)
    permission_id = db.Column(db.Integer, db.ForeignKey("permissions.id", ondelete="CASCADE"), nullable=False)
    role = db.relationship("Role", back_populates="permission_links")
    permission = db.relationship("Permission", back_populates="role_links")

    __table_args__ = (db.UniqueConstraint("role_id", "permission_id", name="uq_role_permission"),)

    def to_dict(self):
        data = super().to_dict()
        data["permission"] = _summary(self.permission, ["id", "module", "action"])
        return data


class BaseUser(TimestampMixin, db.Model):

    __tablename__ = "base_users"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False, unique=True)
    email = db.Column(db.String(150), nullable=False, unique=True, index=True)
    mobile = db.Column(db.String(15), unique=True)
    # Additional contact fields for My Profile ▸ Edit Profile. Not unique —
    # unlike `mobile` there's no uniqueness expectation for a secondary or
    # an emergency contact number.
    other_number = db.Column(db.String(15), nullable=True)
    emergency_contact_number = db.Column(db.String(15), nullable=True)
    password = db.Column(db.String(255), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey("roles.id"), nullable=False, index=True)
    role = db.Column(db.String(20), nullable=False, index=True)
    profile_picture = db.Column(JSONB, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    last_login = db.Column(db.DateTime)
    otp = db.Column(db.String(255), nullable=True)
    otp_expires_at = db.Column(db.DateTime, nullable=True)
    smtp_verified = db.Column(db.Boolean, default=False)
    role_obj = db.relationship("Role", back_populates="users")
    employee = db.relationship("Employee", back_populates="user", uselist=False, cascade="all, delete")
    audit_logs = db.relationship("AuditLog", back_populates="user")

    def __repr__(self):
        return f"<User {self.username}>"

    def to_dict(self):
        data = super().to_dict()
        data.pop("password", None)
        data.pop("otp", None)
        data.pop("otp_expires_at", None)
        data["role_name"] = self.role_obj.name if self.role_obj else self.role
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"]) if self.employee else None
        return data


class Department(TimestampMixin, db.Model):

    __tablename__ = "departments"

    id = db.Column(db.Integer, primary_key=True)
    department_code = db.Column(db.String(20), nullable=False, unique=True)
    department_name = db.Column(db.String(100), nullable=False, unique=True)
    company_id = db.Column(
        db.Integer,
        db.ForeignKey("companies.id"),
        nullable=True,
        index=True,
    )
    branch_id = db.Column(
        db.Integer,
        db.ForeignKey("branches.id"),
        nullable=True,
        index=True,
    )
    description = db.Column(db.Text)
    status = db.Column(db.Boolean, default=True)
    is_active = db.Column(db.Boolean, default=True)
    company = db.relationship("Company", back_populates="departments")
    branch = db.relationship("Branch", back_populates="departments")
    employees = db.relationship("Employee", back_populates="department")
    designations = db.relationship("Designation", back_populates="department")

    def __repr__(self):
        return self.department_name

    def to_dict(self):
        data = super().to_dict()
        data["company"] = (
            {"id": self.company.id, "name": self.company.name, "code": self.company.code}
            if self.company
            else None
        )

        data["branch"] = (
            {"id": self.branch.id, "name": self.branch.name, "code": self.branch.code}
            if self.branch
            else None
        )

        data["employees"] = [
            _summary(employee, ["id", "employee_code", "first_name", "last_name"])
            for employee in self.employees
        ]

        data["designations"] = [
            _summary(designation, ["id", "designation_name", "designation_code"])
            for designation in self.designations
            if designation.is_active
        ]

        return data


class DepartmentHeadcount(TimestampMixin, db.Model):
    __tablename__ = "department_headcounts"

    id = db.Column(db.Integer, primary_key=True)
    department_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=False)
    week_start_date = db.Column(db.Date, nullable=False)
    employee_count = db.Column(db.Integer, nullable=False, default=0)
    updated_by = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    notes = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, default=True)

    department = db.relationship("Department")
    updated_by_employee = db.relationship("Employee", foreign_keys=[updated_by])

    __table_args__ = (
        db.UniqueConstraint("department_id", "week_start_date", name="uq_dept_headcount_week"),
    )

    def to_dict(self):
        data = super().to_dict()
        data["department"] = _summary(self.department, ["id", "department_name", "department_code"])
        data["updated_by_employee"] = _summary(
            self.updated_by_employee, ["id", "employee_code", "first_name", "last_name"]
        )
        return data


class Designation(TimestampMixin, db.Model):

    __tablename__ = "designations"

    id = db.Column(db.Integer,primary_key=True)
    designation_code = db.Column(db.String(20),unique=True)
    designation_name = db.Column(db.String(100),nullable=False)
    department_id = db.Column(
        db.Integer,
        db.ForeignKey(
            "departments.id",
            ondelete="CASCADE"
        ),
        nullable=False
    )
    # Self-referential FK for sub-designations, e.g. IT > Backend > Python/Java.
    # NULL means this is a top-level designation.
    parent_designation_id = db.Column(
        db.Integer,
        db.ForeignKey(
            "designations.id",
            ondelete="CASCADE"
        ),
        nullable=True
    )
    description = db.Column(db.Text)
    status = db.Column(db.Boolean,default=True)
    is_active = db.Column(db.Boolean, default=True)
    is_admin_designation = db.Column(db.Boolean, nullable=False, default=False)
    department = db.relationship("Department",back_populates="designations")
    employees = db.relationship("Employee",back_populates="designation")
    parent_designation = db.relationship(
        "Designation",
        remote_side=[id],
        back_populates="sub_designations"
    )
    sub_designations = db.relationship(
        "Designation",
        back_populates="parent_designation",
        cascade="all, delete-orphan"
    )

    __table_args__ = (

        db.UniqueConstraint(
            "designation_name",
            "department_id",
            name="uq_designation_department"
        ),

    )

    def __repr__(self):
        return self.designation_name

    def to_dict(self):
        data = super().to_dict()
        data["department"] = (
            {
                "id": self.department.id,
                "department_name": self.department.department_name,
                "department_code": self.department.department_code,
                "company": (
                    {"id": self.department.company.id, "name": self.department.company.name}
                    if self.department.company
                    else None
                ),
                "branch": (
                    {"id": self.department.branch.id, "name": self.department.branch.name}
                    if self.department.branch
                    else None
                ),
            }
            if self.department
            else None
        )
        data["employees"] = [
            _summary(employee, ["id", "employee_code", "first_name", "last_name"])
            for employee in self.employees
        ]
        return data


class LeaveType(TimestampMixin, db.Model):

    __tablename__ = "leave_types"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    # "Leave" (multi-day absence) or "Permission" (short in-day absence).
    category = db.Column(db.String(20), nullable=False, default="Leave", server_default="Leave")
    is_active = db.Column(db.Boolean, default=True)
    leaves = db.relationship("Leave", back_populates="leave_type")

    def __repr__(self):
        return f"<LeaveType {self.name}>"

    def to_dict(self):
        data = super().to_dict()
        data["category"] = self.category or "Leave"
        return data


class Holiday(TimestampMixin, db.Model):
    """Single-branch business — no branch_id FK (see implementation_plan.md
    decision log). Organization master, full CRUD."""

    __tablename__ = "holidays"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    holiday_date = db.Column(db.Date, nullable=False)
    holiday_type = db.Column(db.String(20), nullable=False, default="Office", server_default="Office")
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        data = super().to_dict()
        data["holiday_type"] = self.holiday_type or "Office"
        return data

    @classmethod
    def generate_holiday_list_pdf(cls, year=None):
        """Formal holiday-circular PDF, split into two sections:
        Government Holidays and Office Holidays, sorted by date.
        Active holidays only. If `year` is given, filters to that
        calendar year; otherwise includes all active holidays."""
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        query = cls.query.filter(cls.is_active == True)
        if year is not None:
            query = query.filter(db.extract("year", cls.holiday_date) == year)

        holidays = query.order_by(cls.holiday_date).all()
        government = [h for h in holidays if (h.holiday_type or "Office") == "Government"]
        office = [h for h in holidays if (h.holiday_type or "Office") == "Office"]

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "HolidayTitle", parent=styles["Title"], fontSize=18, spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "HolidaySubtitle", parent=styles["Normal"], fontSize=10,
            textColor=colors.HexColor("#64748b"), spaceAfter=20,
        )
        section_style = ParagraphStyle(
            "SectionHeading", parent=styles["Heading2"], fontSize=13,
            textColor=colors.HexColor("#1e293b"), spaceBefore=16, spaceAfter=8,
        )
        empty_style = ParagraphStyle(
            "EmptyNote", parent=styles["Normal"], fontSize=9,
            textColor=colors.HexColor("#94a3b8"), spaceAfter=8,
        )

        elements = []
        heading_text = f"Holiday List{' — ' + str(year) if year else ''}"
        elements.append(Paragraph(heading_text, title_style))
        elements.append(Paragraph("Official schedule of Government and Office holidays", subtitle_style))

        def _build_section(title, rows):
            elements.append(Paragraph(title, section_style))
            if not rows:
                elements.append(Paragraph("No holidays recorded.", empty_style))
                return

            table_data = [["#", "Holiday Name", "Date", "Day"]]
            for index, holiday in enumerate(rows, start=1):
                table_data.append([
                    str(index),
                    holiday.name,
                    holiday.holiday_date.strftime("%d %b %Y"),
                    holiday.holiday_date.strftime("%A"),
                ])

            table = Table(table_data, colWidths=[1.2 * cm, 7 * cm, 3.5 * cm, 3.5 * cm])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(table)

        _build_section("Government Holidays", government)
        elements.append(Spacer(1, 12))
        _build_section("Office Holidays", office)

        doc.build(elements)
        buffer.seek(0)
        return buffer


class Company(TimestampMixin, db.Model):

    __tablename__ = "companies"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False, unique=True)
    code = db.Column(db.String(50), nullable=False, unique=True)
    email = db.Column(db.String(150))
    phone = db.Column(db.String(30))
    website = db.Column(db.String(255))
    address = db.Column(db.Text)
    city = db.Column(db.String(100))
    state = db.Column(db.String(100))
    country = db.Column(db.String(100))
    pincode = db.Column(db.String(20))
    status = db.Column(db.Boolean, default=True)
    is_active = db.Column(db.Boolean, default=True)

    branches = db.relationship(
        "Branch",
        back_populates="company",
        cascade="all, delete-orphan"
    )

    departments = db.relationship(
        "Department",
        back_populates="company",
    )

    def __repr__(self):
        return self.name

    def to_dict(self):
        data = super().to_dict()

        data["branches"] = [
            {
                "id": branch.id,
                "name": branch.name,
                "code": branch.code,
                "email": branch.email,
                "phone": branch.phone,
                "address": branch.address,
                "city": branch.city,
                "state": branch.state,
                "country": branch.country,
                "pincode": branch.pincode,
                "status": branch.status,
                "is_active": branch.is_active,
            }
            for branch in self.branches
            if branch.is_active
        ]

        return data


class Branch(TimestampMixin, db.Model):
    
    __tablename__ = "branches"

    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(
        db.Integer,
        db.ForeignKey("companies.id", ondelete="CASCADE"),
        nullable=False,
        index=True
    )
    name = db.Column(db.String(150), nullable=False)
    code = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(150))
    phone = db.Column(db.String(30))
    address = db.Column(db.Text)
    city = db.Column(db.String(100))
    state = db.Column(db.String(100))
    country = db.Column(db.String(100))
    pincode = db.Column(db.String(20))
    status = db.Column(db.Boolean, default=True)
    is_active = db.Column(db.Boolean, default=True)

    company = db.relationship(
        "Company",
        back_populates="branches"
    )

    departments = db.relationship(
        "Department",
        back_populates="branch",
    )

    __table_args__ = (
        db.UniqueConstraint(
            "company_id",
            "code",
            name="uq_branch_company_code"
        ),
        db.UniqueConstraint(
            "company_id",
            "name",
            name="uq_branch_company_name"
        ),
    )

    def __repr__(self):
        return self.name

    def to_dict(self):
        data = super().to_dict()

        data["company"] = (
            {
                "id": self.company.id,
                "name": self.company.name,
                "code": self.company.code,
            }
            if self.company
            else None
        )

        return data



class Employee(TimestampMixin, db.Model):

    __tablename__ = "employees"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer,db.ForeignKey("base_users.id"),nullable=False,unique=True)
    employee_code = db.Column(db.String(20),unique=True,nullable=False)
    department_id = db.Column(db.Integer,db.ForeignKey("departments.id"),nullable=False)
    designation_id = db.Column(db.Integer,db.ForeignKey("designations.id"),nullable=False)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100))
    gender = db.Column(db.String(20))
    dob = db.Column(db.Date)
    phone = db.Column(db.String(15))
    emergency_contact = db.Column(db.String(15))
    address = db.Column(db.Text)
    city = db.Column(db.String(100))
    state = db.Column(db.String(100))
    country = db.Column(db.String(100))
    pincode = db.Column(db.String(10))
    joining_date = db.Column(db.Date)
    salary = db.Column(db.Numeric(12, 2), default=0)
    allowance = db.Column(db.Numeric(12, 2), default=0)
    pf_number = db.Column(db.String(30))
    esi_number = db.Column(db.String(30))
    account_number = db.Column(db.String(30))
    # IFSC alongside account_number is what a RazorpayX fund account needs
    # to actually pay this employee (see api/v1/crm/razorpay_gateway.py) —
    # optional, so the automated incentive payout keeps working (falling
    # back to an internal settlement) for anyone without it on file yet.
    bank_ifsc = db.Column(db.String(15))
    # Cached once created so a payout doesn't recreate the Razorpay
    # Contact/Fund Account (and doesn't re-hit their API) every month.
    razorpay_contact_id = db.Column(db.String(60))
    razorpay_fund_account_id = db.Column(db.String(60))
    status = db.Column(db.Boolean, default=True)
    is_active = db.Column(db.Boolean, default=True)
    # profile_picture = db.Column(JSONB, nullable=True)
    user = db.relationship("BaseUser", back_populates="employee")
    department = db.relationship("Department",back_populates="employees")
    designation = db.relationship("Designation",back_populates="employees")
    attendances = db.relationship("Attendance",back_populates="employee",cascade="all, delete-orphan")
    manual_attendances = db.relationship("ManualAttendance", back_populates="employee", cascade="all, delete-orphan")
    leaves = db.relationship("Leave",back_populates="employee",cascade="all, delete-orphan")
    network_logs = db.relationship("NetworkStatus",back_populates="employee",cascade="all, delete-orphan")

    def to_dict(self):
        data = super().to_dict()
        data["user"] = _summary(self.user, ["id", "username", "email", "role"]) if self.user else None
        data["department"] = (
            {
                "id": self.department.id,
                "department_name": self.department.department_name,
                "department_code": self.department.department_code,
                "company": (
                    {"id": self.department.company.id, "name": self.department.company.name}
                    if self.department.company
                    else None
                ),
                "branch": (
                    {"id": self.department.branch.id, "name": self.department.branch.name}
                    if self.department.branch
                    else None
                ),
            }
            if self.department
            else None
        )
        data["designation"] = _summary(self.designation, ["id", "designation_name", "designation_code"])
        data["attendance_count"] = len(self.attendances) if self.attendances is not None else 0
        data["leave_count"] = len(self.leaves) if self.leaves is not None else 0
        data["network_log_count"] = len(self.network_logs) if self.network_logs is not None else 0
        return data

    CODE_PREFIX = "ET"

    @classmethod
    def generate_next_code(cls):
        """Next sequential employee code, e.g. ET001, ET002, ... continuing
        from the highest numbered code already in use."""
        prefix = cls.CODE_PREFIX
        max_seq = 0
        codes = db.session.query(cls.employee_code).filter(
            cls.employee_code.like(f"{prefix}%")
        ).all()
        for (code,) in codes:
            suffix = code[len(prefix):]
            if suffix.isdigit():
                max_seq = max(max_seq, int(suffix))
        return f"{prefix}{max_seq + 1:03d}"

    @classmethod
    def generate_employee_report(cls, department_id=None, designation_id=None):
        query = cls.query
        if department_id is not None:
            query = query.filter(cls.department_id == department_id)
        if designation_id is not None:
            query = query.filter(cls.designation_id == designation_id)
        query = query.order_by(cls.id)

        rows = []
        for employee in query.all():
            rows.append([
                employee.employee_code,
                f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
                employee.department.department_name if employee.department else "",
                employee.designation.designation_name if employee.designation else "",
                employee.joining_date.isoformat() if employee.joining_date else "",
                float(employee.salary or 0),
                "Active" if employee.is_active else "Inactive",
            ])
        headers = ["Employee Code", "Name", "Department", "Designation", "Joining Date", "Salary", "Status"]
        return Attendance._create_workbook("Employee Report", headers, rows)


class Attendance(TimestampMixin, db.Model):

    __tablename__ = "attendance"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer,db.ForeignKey("employees.id"),nullable=False)
    attendance_date = db.Column(db.Date,nullable=False)
    check_in = db.Column(db.DateTime)
    check_out = db.Column(db.DateTime)
    working_hours = db.Column(db.Float)
    checkin_latitude = db.Column(db.Numeric(10,7))
    checkin_longitude = db.Column(db.Numeric(10,7))
    checkout_latitude = db.Column(db.Numeric(10,7))
    checkout_longitude = db.Column(db.Numeric(10,7))
    attendance_status = db.Column(db.String(20),default="Present")
    description = db.Column(db.String(255), nullable=True, default="user")
    is_active = db.Column(db.Boolean, default=True)

    # ---- Advanced check-in / permission / overtime workflow ----
    # `check_in` / `check_out` above stay as the FIRST check-in and LAST
    # check-out of the day (kept for list views + backward compatibility);
    # the full timeline lives in `events`. Everything below is derived by
    # attendance_engine.recompute_attendance().
    gross_working_hours = db.Column(db.Float, default=0)      # sum of in->out segments, before breaks
    late_login_minutes = db.Column(db.Float, default=0)
    late_login_reason = db.Column(db.String(255), nullable=True)
    permission_minutes = db.Column(db.Float, default=0)       # sum of out->in gaps
    permission_over_limit = db.Column(db.Boolean, default=False)
    nap_minutes = db.Column(db.Float, default=0)              # snapshot of AttendanceSetting at compute time
    lunch_minutes = db.Column(db.Float, default=0)
    tea_minutes = db.Column(db.Float, default=0)
    overtime_hours = db.Column(db.Float, default=0)           # working_hours beyond required
    overtime_reason = db.Column(db.String(255), nullable=True)

    employee = db.relationship("Employee",back_populates="attendances")
    events = db.relationship(
        "AttendanceEvent",
        back_populates="attendance",
        cascade="all, delete-orphan",
        order_by="AttendanceEvent.event_time",
    )

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        data["events"] = [
            e.to_dict() for e in sorted(
                (self.events or []),
                key=lambda ev: (ev.event_time or datetime.min),
            )
            if e.is_active
        ]
        return data

    @classmethod
    def _create_workbook(cls, title, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        thin = Side(border_style="thin", color="000000")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        centered = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value="HRMS").font = Font(size=18, bold=True)
        ws.cell(row=1, column=1).alignment = centered

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value=title).font = Font(size=14, bold=True)
        ws.cell(row=2, column=1).alignment = centered

        header_row = 4
        for index, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_index, row_data in enumerate(rows, start=header_row + 1):
            for col_index, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        column_widths = [18, 30, 18, 16, 16, 16, 16]
        for index, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=header_row, column=index).column_letter].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def generate_attendance_report(cls, from_date=None, to_date=None, employee_id=None):
        query = cls.query.join(Employee)
        if employee_id is not None:
            query = query.filter(cls.employee_id == employee_id)
        if from_date and to_date:
            query = query.filter(cls.attendance_date.between(from_date, to_date))
        elif from_date:
            query = query.filter(cls.attendance_date >= from_date)
        elif to_date:
            query = query.filter(cls.attendance_date <= to_date)

        query = query.order_by(cls.attendance_date, cls.employee_id)

        rows = []
        for attendance in query.all():
            rows.append([
                attendance.employee_id,
                f"{attendance.employee.first_name or ''} {attendance.employee.last_name or ''}".strip(),
                attendance.attendance_date.isoformat(),
                attendance.check_in.strftime("%H:%M") if attendance.check_in else "",
                attendance.check_out.strftime("%H:%M") if attendance.check_out else "",
            ])

        headers = ["Employee ID", "Employee Name", "Date", "Check-in Time", "Check-out Time"]
        return cls._create_workbook("Attendance Report", headers, rows)

    @classmethod
    def get_monthly_summary(cls, employee_id, month, year):
        """Aggregates a single employee's attendance for one calendar month
        into present/absent/leave/holiday counts and a salary calculation.
        Holiday dates (Government or Office, active only) are excluded both
        from the deductible-absent set and from the payable-days divisor —
        the same treatment approved leave already gets."""
        from calendar import monthrange

        employee = Employee.query.get(employee_id)
        if not employee:
            return None

        days_in_month = monthrange(year, month)[1]
        from_date = date(year, month, 1)
        to_date = date(year, month, days_in_month)

        records = cls.query.filter(
            cls.employee_id == employee_id,
            cls.attendance_date.between(from_date, to_date),
        ).all()

        present_days = sum(1 for r in records if r.attendance_status == "Present")
        marked_absent_dates = {r.attendance_date for r in records if r.attendance_status == "Absent"}

        leave_query = Leave.query.filter(
            Leave.employee_id == employee_id,
            Leave.status == "Approved",
            Leave.is_active == True,
            Leave.from_date <= to_date,
            Leave.to_date >= from_date,
        )
        approved_leave_dates = set()
        for leave in leave_query.all():
            day = max(from_date, leave.from_date)
            end = min(to_date, leave.to_date)
            while day <= end:
                approved_leave_dates.add(day)
                day += timedelta(days=1)

        holiday_query = Holiday.query.filter(
            Holiday.is_active == True,
            Holiday.holiday_date.between(from_date, to_date),
        )
        holiday_dates = {h.holiday_date for h in holiday_query.all()}

        deductible_absent_dates = marked_absent_dates - approved_leave_dates - holiday_dates
        total_worked_hours = sum(r.working_hours or 0 for r in records)

        salary_value = float(employee.salary or 0)
        allowance_value = float(employee.allowance or 0)
        payable_days = days_in_month - len(holiday_dates)
        per_day_salary = (salary_value / payable_days) if payable_days else 0.0

        # Approved leave and holidays are PAID — only genuine unexcused
        # absences reduce pay.
        leave_deduction = 0.0
        absent_deduction = round(per_day_salary * len(deductible_absent_dates), 2)
        total_deduction = round(leave_deduction + absent_deduction, 2)

        gross_salary = round(salary_value + allowance_value, 2)

        # CRM-department employees also earn a monthly incentive
        # (crm/employee_incentives). It is additional pay, so it is shown
        # here and added on top of the post-deduction salary. For every
        # other department the incentive is not applicable -> None.
        department_name = (
            (getattr(employee.department, "department_name", "") or "")
            .strip()
            .lower()
        )
        is_crm = department_name == "crm"

        incentive_amount = None
        incentive_source = None
        incentive_registrations = None
        incentive_eligible = None
        # Invoice details for the incentive — surfaced alongside the
        # amount so Finance can see whether it's actually been invoiced
        # (and paid) rather than just a calculated figure.
        incentive_invoice_number = None
        incentive_invoice_status = None
        incentive_invoice_due_date = None
        incentive_invoice_paid_amount = None
        if is_crm:
            # Computed LIVE from this month's actual Registration
            # (Meeting) rows every time this runs, rather than reading a
            # MonthlyPayout snapshot that's only refreshed when the CRM
            # Incentives screen's "Run" button (or the automated 20th-of-
            # next-month payout) has executed for this exact month — that
            # snapshot going stale is why Finance was seeing ₹0.00 /
            # "0 eligible / 0 reg" for CRM employees who clearly had
            # registrations. Same flat rule as incentive_engine.py:
            # Rs.1,000 for the first 30 registrations in the month, +0.6%
            # of Rs.1,000 (Rs.6) per registration from the 31st onward.
            days_in_month_local = monthrange(year, month)[1]
            month_start = datetime(year, month, 1)
            month_end = datetime(year, month, days_in_month_local) + timedelta(days=1)

            incentive_registrations = (
                db.session.query(db.func.count(Meeting.id))
                .filter(
                    Meeting.registered_by == employee.id,
                    Meeting.is_active == True,
                    Meeting.created_at >= month_start,
                    Meeting.created_at < month_end,
                )
                .scalar()
                or 0
            )

            target_count = 30
            incentive_eligible = max(0, incentive_registrations - target_count)
            if incentive_registrations > 0:
                extra_rate = 0.006 * 1000.0  # 0.6% of Rs.1,000 per registration past 30
                incentive_amount = round(1000.0 + incentive_eligible * extra_rate, 2)
                incentive_source = "tier"
            else:
                incentive_amount = 0.0
                incentive_source = "none"

            # Invoice/payment details are still read from whatever the CRM
            # incentive engine has actually generated for this month (the
            # amount above no longer depends on it existing) so Finance can
            # see real invoice/payment status when one has been raised.
            payout = MonthlyPayout.query.filter(
                MonthlyPayout.employee_id == employee.id,
                MonthlyPayout.month == month,
                MonthlyPayout.year == year,
                MonthlyPayout.is_active == True,
            ).first()

            if payout is not None:
                invoice = Invoice.query.filter(
                    Invoice.monthly_payout_id == payout.id,
                    Invoice.is_active == True,
                ).order_by(Invoice.id.desc()).first()

                if invoice is not None:
                    incentive_invoice_number = invoice.invoice_number
                    incentive_invoice_status = invoice.status
                    incentive_invoice_due_date = (
                        invoice.due_date.isoformat() if invoice.due_date else None
                    )
                    incentive_invoice_paid_amount = round(
                        float(sum((p.amount or 0) for p in invoice.payments)), 2
                    )

            if incentive_invoice_number is None:
                # No legacy fallback needed once live computation covers
                # everything, but EmployeeIncentive rows (pre-flat-rule
                # data) still count toward "has this been invoiced before".
                incentive_rows = EmployeeIncentive.query.filter(
                    EmployeeIncentive.employee_id == employee.id,
                    EmployeeIncentive.month == month,
                    EmployeeIncentive.year == year,
                    EmployeeIncentive.is_active == True,
                ).all()
                if incentive_rows and incentive_amount == 0.0:
                    incentive_amount = round(
                        sum(float(r.calculated_amount or 0) for r in incentive_rows), 2
                    )
                    incentive_source = "slab"

        net_salary = round(
            gross_salary - total_deduction + (incentive_amount or 0.0), 2
        )

        return {
            "employee_id": employee.id,
            "employee_code": employee.employee_code,
            "employee_name": f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
            "department": getattr(employee.department, "department_name", None),
            "is_crm": is_crm,
            "month": month,
            "year": year,
            "days_in_month": days_in_month,
            "holiday_days": len(holiday_dates),
            "payable_days": payable_days,
            "present_days": present_days,
            "absent_days": len(deductible_absent_dates),
            "approved_leave_days": len(approved_leave_dates),
            "worked_hours": round(total_worked_hours, 2),
            "salary": salary_value,
            "allowance": allowance_value,
            "gross_salary": gross_salary,
            "per_day_salary": round(per_day_salary, 2),
            "leave_deduction": leave_deduction,
            "absent_deduction": absent_deduction,
            "total_deduction": total_deduction,
            "incentive_amount": incentive_amount,
            "incentive_source": incentive_source,
            "incentive_registrations": incentive_registrations,
            "incentive_eligible": incentive_eligible,
            "incentive_invoice_number": incentive_invoice_number,
            "incentive_invoice_status": incentive_invoice_status,
            "incentive_invoice_due_date": incentive_invoice_due_date,
            "incentive_invoice_paid_amount": incentive_invoice_paid_amount,
            "net_salary": net_salary,
        }

    @classmethod
    def get_monthly_summary_list(cls, month, year, employee_id=None):
        query = Employee.query
        if employee_id is not None:
            query = query.filter(Employee.id == employee_id)
        return [
            cls.get_monthly_summary(e.id, month, year)
            for e in query.order_by(Employee.id).all()
        ]

    @classmethod
    def generate_salary_report(
        cls,
        from_date=None,
        to_date=None,
        employee_id=None
    ):

        query = Employee.query

        if employee_id is not None:

            query = query.filter(
                Employee.id == employee_id
            )

        rows = []

        for employee in query.order_by(
            Employee.id
        ).all():


            attendance_query = cls.query.filter(
                cls.employee_id == employee.id
            )

            if from_date and to_date:

                attendance_query = attendance_query.filter(
                    cls.attendance_date.between(
                        from_date,
                        to_date
                    )
                )

            elif from_date:

                attendance_query = attendance_query.filter(
                    cls.attendance_date >= from_date
                )

            elif to_date:

                attendance_query = attendance_query.filter(
                    cls.attendance_date <= to_date
                )

            attendance_records = attendance_query.all()

            attendance_count = len(
                attendance_records
            )

            salary_value = float(
                employee.salary or 0
            )

            allowance_value = float(
                employee.allowance or 0
            )


            if from_date and to_date:

                period_start = from_date
                period_end = to_date

                days_in_period = (
                    to_date - from_date
                ).days + 1

            elif from_date:

                period_start = from_date
                period_end = date.today()

                days_in_period = (
                    period_end - period_start
                ).days + 1

            elif to_date:

                period_end = to_date
                period_start = to_date.replace(
                    day=1
                )

                days_in_period = (
                    period_end - period_start
                ).days + 1

            else:

                period_start = None
                period_end = None
                days_in_period = 30

        

            leave_query = Leave.query.filter(
                Leave.employee_id == employee.id,
                Leave.status == "Approved",
                Leave.is_active == True,
            )

            if period_start and period_end:

                leave_query = leave_query.filter(
                    Leave.from_date <= period_end,
                    Leave.to_date >= period_start
                )

            elif period_start:

                leave_query = leave_query.filter(
                    Leave.to_date >= period_start
                )

            elif period_end:

                leave_query = leave_query.filter(
                    Leave.from_date <= period_end
                )

            approved_leave_dates = set()

            for leave in leave_query.all():

                leave_start = (
                    max(
                        period_start,
                        leave.from_date
                    )
                    if period_start
                    else leave.from_date
                )

                leave_end = (
                    min(
                        period_end,
                        leave.to_date
                    )
                    if period_end
                    else leave.to_date
                )

                day = leave_start

                while day <= leave_end:

                    approved_leave_dates.add(day)

                    day += timedelta(days=1)


            holiday_dates = set()

            holiday_query = Holiday.query.filter(
                Holiday.is_active == True
            )

            if period_start and period_end:

                holiday_query = holiday_query.filter(
                    Holiday.holiday_date.between(
                        period_start,
                        period_end
                    )
                )

            elif period_start:

                holiday_query = holiday_query.filter(
                    Holiday.holiday_date >= period_start
                )

            elif period_end:

                holiday_query = holiday_query.filter(
                    Holiday.holiday_date <= period_end
                )

            for holiday in holiday_query.all():

                holiday_dates.add(
                    holiday.holiday_date
                )

            marked_absent_dates = {
                record.attendance_date
                for record in attendance_records
                if record.attendance_status == "Absent"
            }


            deductible_absent_dates = (
                marked_absent_dates
                - approved_leave_dates
                - holiday_dates
            )


            payable_days = (
                days_in_period
                - len(holiday_dates)
            )

            if payable_days < 0:
                payable_days = 0


            per_day_salary = (
                salary_value / payable_days
                if payable_days
                else 0.0
            )


            absent_deduction = round(
                per_day_salary
                * len(deductible_absent_dates),
                2
            )


            gross_salary = round(
                salary_value + allowance_value,
                2
            )


            net_salary = round(
                gross_salary - absent_deduction,
                2
            )


            approved_leave_dates_label = ", ".join(
                day.isoformat()
                for day in sorted(
                    approved_leave_dates
                )
            )


            holiday_dates_label = ", ".join(
                day.isoformat()
                for day in sorted(
                    holiday_dates
                )
            )


            rows.append([

                employee.id,

                employee.employee_code,

                (
                    f"{employee.first_name or ''} "
                    f"{employee.last_name or ''}"
                ).strip(),

                salary_value,

                attendance_count,

                len(deductible_absent_dates),

                len(approved_leave_dates),

                approved_leave_dates_label,

                absent_deduction,

                net_salary,

                from_date.isoformat()
                if from_date
                else "",

                to_date.isoformat()
                if to_date
                else "",
            ])


        headers = [

            "Employee ID",

            "Employee Code",

            "Employee Name",

            "Salary",

            "Attendance Count",

            "Absent Days (Deducted)",

            "Approved Leave Days",

            "Approved Leave Dates",

            "Absent Deduction",

            "Net Salary",

            "From Date",

            "To Date",
        ]

        return cls._create_workbook(
            "Salary Report",
            headers,
            rows
        )


class ManualAttendance(TimestampMixin, db.Model):

    __tablename__ = "manual_attendance"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    attendance_date = db.Column(db.Date, nullable=False)
    check_in = db.Column(db.DateTime)
    check_out = db.Column(db.DateTime)
    working_hours = db.Column(db.Float)
    checkin_latitude = db.Column(db.Numeric(10,7))
    checkin_longitude = db.Column(db.Numeric(10,7))
    checkout_latitude = db.Column(db.Numeric(10,7))
    checkout_longitude = db.Column(db.Numeric(10,7))
    attendance_status = db.Column(db.String(20), default="Present")
    description = db.Column(db.String(255), nullable=True, default="admin")
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee", back_populates="manual_attendances")

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data


class Leave(TimestampMixin, db.Model):

    __tablename__ = "leave"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    leave_type_id = db.Column(db.Integer, db.ForeignKey("leave_types.id"), nullable=False)
    from_date = db.Column(db.Date)
    to_date = db.Column(db.Date)
    total_days = db.Column(db.Integer)
    reason = db.Column(db.Text)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default="Pending")
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee", back_populates="leaves")
    leave_type = db.relationship("LeaveType", back_populates="leaves")

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        data["leave_type"] = _summary(self.leave_type, ["id", "name", "category"]) if self.leave_type else None
        return data

    @classmethod
    def generate_leave_report(cls, from_date=None, to_date=None, employee_id=None):
        query = cls.query.join(Employee)
        if employee_id is not None:
            query = query.filter(cls.employee_id == employee_id)
        if from_date and to_date:
            query = query.filter(cls.from_date <= to_date, cls.to_date >= from_date)
        elif from_date:
            query = query.filter(cls.to_date >= from_date)
        elif to_date:
            query = query.filter(cls.from_date <= to_date)
        query = query.order_by(cls.from_date, cls.employee_id)

        rows = []
        for leave in query.all():
            rows.append([
                leave.employee_id,
                f"{leave.employee.first_name or ''} {leave.employee.last_name or ''}".strip(),
                leave.leave_type.name if leave.leave_type else "",
                leave.from_date.isoformat() if leave.from_date else "",
                leave.to_date.isoformat() if leave.to_date else "",
                leave.total_days or 0,
                leave.status,
            ])
        headers = ["Employee ID", "Employee Name", "Leave Type", "From Date", "To Date", "Total Days", "Status"]
        return Attendance._create_workbook("Leave Report", headers, rows)


class NetworkStatus(TimestampMixin, db.Model):

    __tablename__ = "network_status"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer,db.ForeignKey("employees.id"),nullable=False)
    latitude = db.Column(db.Numeric(10,7))
    longitude = db.Column(db.Numeric(10,7))
    ip_address = db.Column(db.String(50))
    network_type = db.Column(db.String(50))
    device_name = db.Column(db.String(100))
    battery_percentage = db.Column(db.Integer)
    is_online = db.Column(db.Boolean,default=True)
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee",back_populates="network_logs")

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data


class AuditLog(TimestampMixin, db.Model):

    __tablename__ = "audit_logs"

    id = db.Column(db.Integer,primary_key=True)
    user_id = db.Column(db.Integer,db.ForeignKey("base_users.id"),nullable=False)
    action = db.Column(db.String(100),nullable=False)
    table_name = db.Column(db.String(100))
    record_id = db.Column(db.Integer)
    description = db.Column(db.Text)
    ip_address = db.Column(db.String(50))
    is_active = db.Column(db.Boolean, default=True)
    user = db.relationship("BaseUser",back_populates="audit_logs")

    def to_dict(self):
        data = super().to_dict()
        data["user"] = _summary(self.user, ["id", "username", "email", "role"])
        return data


# ===========================================================================
# Phase 2 — Employee lifecycle (add-only lists; soft-delete via is_active)
# ===========================================================================

class EmployeeDocument(TimestampMixin, db.Model):
    __tablename__ = "employee_documents"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    doc_type = db.Column(db.String(50), nullable=False)
    file_url = db.Column(db.String(255), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee")

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data


class EmployeePermission(TimestampMixin, db.Model):
    """Short-leave / gate-pass request — distinct from Role/RolePermission."""
    __tablename__ = "employee_permissions"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    permission_date = db.Column(db.Date, nullable=False)
    from_time = db.Column(db.Time, nullable=False)
    to_time = db.Column(db.Time, nullable=False)
    reason = db.Column(db.Text)
    status = db.Column(db.String(20), default="Pending")
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee")

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data


class Overtime(TimestampMixin, db.Model):
    __tablename__ = "overtime"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    overtime_date = db.Column(db.Date, nullable=False)
    hours = db.Column(db.Numeric(4, 2), nullable=False)
    description = db.Column(db.Text,nullable=True)
    status = db.Column(db.String(20), default="Pending")
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee")

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        data["description"] = (self.description)
        return data


class Payroll(TimestampMixin, db.Model):
    __tablename__ = "payroll"

    id = db.Column(
        db.Integer,
        primary_key=True,
    )

    employee_id = db.Column(
        db.Integer,
        db.ForeignKey("employees.id"),
        nullable=False,
    )

    pay_month = db.Column(
        db.String(7),
        nullable=False,
    )

    gross_salary = db.Column(
        db.Numeric(12, 2),
        default=0,
    )

    deductions = db.Column(
        db.Numeric(12, 2),
        default=0,
    )

    net_salary = db.Column(
        db.Numeric(12, 2),
        default=0,
    )

    status = db.Column(
        db.String(20),
        default="Draft",
    )

    is_active = db.Column(
        db.Boolean,
        default=True,
    )

    employee = db.relationship(
        "Employee"
    )

    def to_dict(self):
        data = super().to_dict()

        employee = self.employee

        if employee is None:
            data["employee"] = None
            data["department_id"] = None
            data["designation_id"] = None
            data["branch_id"] = None
            data["company_id"] = None
            return data

        employee_data = {
            "id": employee.id,
            "employee_code": employee.employee_code,
            "first_name": employee.first_name,
            "last_name": employee.last_name,
        }

        department = getattr(
            employee,
            "department",
            None,
        )

        if department is not None:
            department_data = {
                "id": department.id,
                "department_name": department.department_name,
                "department_code": department.department_code,
            }

            company = getattr(
                department,
                "company",
                None,
            )

            if company is not None:
                department_data["company"] = {
                    "id": company.id,
                    "name": getattr(
                        company,
                        "name",
                        None,
                    ),
                }
            else:
                department_data["company"] = None

            branch = getattr(
                department,
                "branch",
                None,
            )

            if branch is not None:
                department_data["branch"] = {
                    "id": branch.id,
                    "name": getattr(
                        branch,
                        "name",
                        None,
                    ),
                }
            else:
                department_data["branch"] = None

            employee_data["department"] = (
                department_data
            )
        else:
            employee_data["department"] = None

        designation = getattr(
            employee,
            "designation",
            None,
        )

        if designation is not None:
            employee_data["designation"] = {
                "id": designation.id,
                "designation_name": designation.designation_name,
                "designation_code": designation.designation_code,
            }
        else:
            employee_data["designation"] = None

        data["employee"] = employee_data

        data["department_id"] = (
            employee.department_id
        )

        data["designation_id"] = (
            employee.designation_id
        )

        if department is not None:
            branch = getattr(
                department,
                "branch",
                None,
            )

            company = getattr(
                department,
                "company",
                None,
            )

            data["branch_id"] = (
                branch.id
                if branch is not None
                else None
            )

            data["company_id"] = (
                company.id
                if company is not None
                else None
            )
        else:
            data["branch_id"] = None
            data["company_id"] = None

        return data

    @classmethod
    def generate_payroll_report(cls, from_month=None, to_month=None, employee_id=None):
        query = cls.query.join(Employee)
        if employee_id is not None:
            query = query.filter(cls.employee_id == employee_id)
        if from_month:
            query = query.filter(cls.pay_month >= from_month)
        if to_month:
            query = query.filter(cls.pay_month <= to_month)
        query = query.order_by(cls.pay_month, cls.employee_id)

        rows = []
        for payroll in query.all():
            rows.append([
                payroll.employee_id,
                f"{payroll.employee.first_name or ''} {payroll.employee.last_name or ''}".strip(),
                payroll.pay_month,
                float(payroll.gross_salary or 0),
                float(payroll.deductions or 0),
                float(payroll.net_salary or 0),
                payroll.status,
            ])
        headers = ["Employee ID", "Employee Name", "Pay Month", "Gross Salary", "Deductions", "Net Salary", "Status"]
        return Attendance._create_workbook("Payroll Report", headers, rows)


class Performance(TimestampMixin, db.Model):
    __tablename__ = "performance_reviews"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    company_id = db.Column( db.Integer, db.ForeignKey("companies.id"), nullable=True, index=True, )
    branch_id = db.Column( db.Integer, db.ForeignKey("branches.id"), nullable=True, index=True, )
    department_id = db.Column( db.Integer, db.ForeignKey("departments.id"), nullable=True, index=True, )
    designation_id = db.Column( db.Integer, db.ForeignKey("designations.id"), nullable=True, index=True, )
    review_period = db.Column( db.String(20), nullable=False, index=True, )
    day_to_day_performance = db.Column( db.Numeric(3, 1), nullable=True, )
    work_performance = db.Column( db.Numeric(3, 1), nullable=True, )
    behavioral_performance = db.Column( db.Numeric(3, 1), nullable=True, )
    rating = db.Column( db.Numeric(3, 1), nullable=True, )
    remarks = db.Column( db.Text, nullable=True, )
    is_active = db.Column( db.Boolean, default=True, nullable=False, index=True, )
    employee = db.relationship( "Employee", foreign_keys=[employee_id], )
    company = db.relationship( "Company", foreign_keys=[company_id], )
    branch = db.relationship( "Branch", foreign_keys=[branch_id], )
    department = db.relationship( "Department", foreign_keys=[department_id], )
    designation = db.relationship( "Designation", foreign_keys=[designation_id], )

    def to_dict(self): 
      data = super().to_dict()
      data["employee"] = _summary( self.employee, [ "id", "employee_code", "first_name", "last_name", ], ) 
      data["company"] = _summary( self.company, [ "id", "name", ], )
      data["branch"] = _summary( self.branch, [ "id", "name", ], ) 
      data["department"] = _summary( self.department, [ "id", "department_name", "department_code", ], ) 
      data["designation"] = _summary( self.designation, [ "id", "designation_name", "designation_code", ], ) 
      return data


class Training(TimestampMixin, db.Model):
    __tablename__ = "trainings"
 
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    program_name = db.Column(db.String(150), nullable=False)
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    status = db.Column(db.String(20), default="Scheduled")
    status_description = db.Column(db.Text)
    performance = db.Column(db.String(20), default="Not Rated")
    performance_description = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)
    employee = db.relationship("Employee")
 
    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data


class Promotion(TimestampMixin, db.Model):
    __tablename__ = "promotions"

    id = db.Column(
        db.Integer,
        primary_key=True,
    )

    employee_id = db.Column(
        db.Integer,
        db.ForeignKey("employees.id"),
        nullable=False,
    )

    from_designation_id = db.Column(
        db.Integer,
        db.ForeignKey("designations.id"),
        nullable=False,
    )

    to_designation_id = db.Column(
        db.Integer,
        db.ForeignKey("designations.id"),
        nullable=False,
    )

    promotion_date = db.Column(
        db.Date,
        nullable=False,
    )

    reason = db.Column(
        db.String(255),
        nullable=False,
        default="Other",
    )

    accomplishments = db.Column(
        db.Text,
        nullable=True,
    )

    is_active = db.Column(
        db.Boolean,
        default=True,
        nullable=False,
    )

    employee = db.relationship(
        "Employee",
        foreign_keys=[employee_id],
    )

    from_designation = db.relationship(
        "Designation",
        foreign_keys=[from_designation_id],
    )

    to_designation = db.relationship(
        "Designation",
        foreign_keys=[to_designation_id],
    )

    def to_dict(self):
        data = super().to_dict()

        data["promotion_date"] = (
            self.promotion_date
        )

        data["reason"] = (
            self.reason or "Other"
        )

        data["accomplishments"] = (
            self.accomplishments or ""
        )

        data["employee"] = _summary(
            self.employee,
            [
                "id",
                "employee_code",
                "first_name",
                "last_name",
            ],
        )

        data["from_designation"] = _summary(
            self.from_designation,
            [
                "id",
                "designation_name",
            ],
        )

        data["to_designation"] = _summary(
            self.to_designation,
            [
                "id",
                "designation_name",
            ],
        )

        return data


class Transfer(TimestampMixin, db.Model):
    __tablename__ = "transfers"

    id = db.Column(
        db.Integer,
        primary_key=True,
    )

    employee_id = db.Column(
        db.Integer,
        db.ForeignKey("employees.id"),
        nullable=False,
    )

    from_department_id = db.Column(
        db.Integer,
        db.ForeignKey("departments.id"),
        nullable=True,
    )

    to_department_id = db.Column(
        db.Integer,
        db.ForeignKey("departments.id"),
        nullable=False,
    )

    transfer_reason = db.Column(
        db.String(255),
        nullable=False,
        default="Other",
    )

    transfer_apply_date = db.Column(
        db.Date,
        nullable=False,
    )

    relieving_date = db.Column(
        db.Date,
        nullable=False,
    )

    joining_date = db.Column(
        db.Date,
        nullable=False,
    )

    # Transfer destination / new location
    location = db.Column(
        db.String(255),
        nullable=True,
    )

    accomplishments = db.Column(
        db.Text,
        nullable=True,
    )

    is_active = db.Column(
        db.Boolean,
        default=True,
        nullable=False,
    )

    employee = db.relationship(
        "Employee",
        foreign_keys=[employee_id],
    )

    from_department = db.relationship(
        "Department",
        foreign_keys=[from_department_id],
    )

    to_department = db.relationship(
        "Department",
        foreign_keys=[to_department_id],
    )

    def to_dict(self):
        data = super().to_dict()

        data["transfer_reason"] = (
            self.transfer_reason or "Other"
        )

        data["transfer_apply_date"] = (
            self.transfer_apply_date
        )

        data["relieving_date"] = (
            self.relieving_date
        )

        data["joining_date"] = (
            self.joining_date
        )

        data["location"] = (
            self.location or ""
        )

        data["accomplishments"] = (
            self.accomplishments or ""
        )

        employee = self.employee

        if employee:
            data["employee"] = _summary(
                employee,
                [
                    "id",
                    "employee_code",
                    "first_name",
                    "last_name",
                    "city",
                    "state",
                    "country",
                    "address",
                ],
            )


            employee_location = getattr(
                employee,
                "location",
                None,
            )

            if (
                isinstance(
                    employee_location,
                    str,
                )
                and employee_location.strip()
            ):
                data["current_location"] = (
                    employee_location.strip()
                )

            else:
                location_parts = [
                    getattr(
                        employee,
                        "city",
                        None,
                    ),
                    getattr(
                        employee,
                        "state",
                        None,
                    ),
                    getattr(
                        employee,
                        "country",
                        None,
                    ),
                ]

                location_parts = [
                    str(value).strip()
                    for value in location_parts
                    if value is not None
                    and str(value).strip()
                ]

                if location_parts:
                    data["current_location"] = (
                        ", ".join(
                            location_parts
                        )
                    )

                elif (
                    getattr(
                        employee,
                        "address",
                        None,
                    )
                    and str(
                        employee.address
                    ).strip()
                ):
                    data["current_location"] = (
                        str(
                            employee.address
                        ).strip()
                    )

                else:
                    data["current_location"] = ""

            data["employee_id"] = (
                employee.id
            )

            data["employee_code"] = (
                employee.employee_code
            )

        else:
            data["employee"] = None
            data["current_location"] = ""
            data["employee_id"] = (
                self.employee_id
            )
            data["employee_code"] = ""

        if self.from_department:
            data["from_department"] = _summary(
                self.from_department,
                [
                    "id",
                    "department_name",
                    "department_code",
                ],
            )
        else:
            data["from_department"] = None

        data["from_department_id"] = (
            self.from_department_id
        )

        if self.to_department:
            data["to_department"] = _summary(
                self.to_department,
                [
                    "id",
                    "department_name",
                    "department_code",
                ],
            )
        else:
            data["to_department"] = None

        data["to_department_id"] = (
            self.to_department_id
        )


        data["is_active"] = (
            self.is_active
        )

        return data


class Resignation(TimestampMixin, db.Model):
    __tablename__ = "resignations"

    id = db.Column(db.Integer, primary_key=True)

    employee_id = db.Column(
        db.Integer,
        db.ForeignKey("employees.id"),
        nullable=False,
    )

    notice_date = db.Column(db.Date, nullable=False)
    last_working_date = db.Column(db.Date, nullable=False)

    reason = db.Column(db.Text)
    description = db.Column(db.Text)
    accomplishments = db.Column(db.Text)

    # Previous organization details
    previous_company_id = db.Column(
        db.Integer,
        db.ForeignKey("companies.id"),
        nullable=   True,
    )

    previous_branch_id = db.Column(
        db.Integer,
        db.ForeignKey("branches.id"),
        nullable=True,
    )

    previous_department_id = db.Column(
        db.Integer,
        db.ForeignKey("departments.id"),
        nullable=True,
    )

    previous_designation_id = db.Column(
        db.Integer,
        db.ForeignKey("designations.id"),
        nullable=True,
    )

    status = db.Column(
        db.String(20),
        default="Pending",
    )

    is_active = db.Column(
        db.Boolean,
        default=True,
    )

    # Employee
    employee = db.relationship(
        "Employee",
        foreign_keys=[employee_id],
    )

    # Previous organization
    previous_company = db.relationship(
        "Company",
        foreign_keys=[previous_company_id],
    )

    previous_branch = db.relationship(
        "Branch",
        foreign_keys=[previous_branch_id],
    )

    previous_department = db.relationship(
        "Department",
        foreign_keys=[previous_department_id],
    )

    previous_designation = db.relationship(
        "Designation",
        foreign_keys=[previous_designation_id],
    )

    def to_dict(self):
        data = super().to_dict()

        data["employee"] = _summary(
            self.employee,
            [
                "id",
                "employee_code",
                "first_name",
                "last_name",
            ],
        )

        data["previous_organization"] = {
            "company": _summary(
                self.previous_company,
                [
                    "id",
                    "name",
                ],
            ),
            "branch": _summary(
                self.previous_branch,
                [
                    "id",
                    "name",
                ],
            ),
            "department": _summary(
                self.previous_department,
                [
                    "id",
                    "department_name",
                ],
            ),
            "designation": _summary(
                self.previous_designation,
                [
                    "id",
                    "designation_name",
                ],
            ),
        }

        return data


class ExitManagement(TimestampMixin, db.Model):
    __tablename__ = "exit_management"

    id = db.Column(db.Integer, primary_key=True)
    resignation_id = db.Column(db.Integer, db.ForeignKey("resignations.id"), nullable=False)
    clearance_status = db.Column(db.String(20), default="Pending")
    exit_date = db.Column(db.Date)
    remarks = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)
    resignation = db.relationship("Resignation")

    def to_dict(self):
        data = super().to_dict()
        data["resignation"] = _summary(self.resignation, ["id", "employee_id", "status", "last_working_date"])
        return data


# ===========================================================================
# Phase 3 — CRM
# ===========================================================================

class Lead(TimestampMixin, db.Model):
    __tablename__ = "leads"

    id = db.Column(db.Integer, primary_key=True)
    lead_name = db.Column(db.String(150), nullable=False)
    contact_number = db.Column(db.String(15))
    email = db.Column(db.String(150))
    source = db.Column(db.String(50))
    status = db.Column(db.String(20), default="New")
    # Free-text notes — for a photo/OCR-created lead this holds the raw
    # extracted text so a human can verify/correct anything the OCR
    # engine misread.
    notes = db.Column(db.Text, nullable=True)
    assigned_to = db.Column(db.Integer, db.ForeignKey("employees.id"))
    created_by = db.Column(db.Integer, db.ForeignKey("employees.id"))
    upload_batch_id = db.Column(db.Integer, db.ForeignKey("lead_upload_batches.id"), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    assignee = db.relationship("Employee", foreign_keys=[assigned_to])
    creator = db.relationship("Employee", foreign_keys=[created_by])
    upload_batch = db.relationship("LeadUploadBatch", back_populates="leads")
    customers = db.relationship("Customer", back_populates="lead")

    @staticmethod
    def _derive_team_type(department_name):
        if not department_name:
            return None
        normalized = department_name.strip().lower()
        if "non" in normalized and "voice" in normalized:
            return "Non-Voice"
        if "voice" in normalized:
            return "Voice"
        return None

    @classmethod
    def _employee_hierarchy(cls, employee):
        if not employee:
            return None

        department = getattr(employee, "department", None)
        designation = getattr(employee, "designation", None)
        company = getattr(department, "company", None) if department else None
        branch = getattr(department, "branch", None) if department else None
        department_name = getattr(department, "department_name", None)

        return {
            "company": _summary(company, ["id", "name"]),
            "branch": _summary(branch, ["id", "name"]),
            "department": _summary(department, ["id", "department_name"]),
            "designation": _summary(designation, ["id", "designation_name"]),
            "team_type": cls._derive_team_type(department_name),
        }

    def to_dict(self):
        data = super().to_dict()
        data["assignee"] = _summary(self.assignee, ["id", "employee_code", "first_name", "last_name"])
        data["creator"] = _summary(self.creator, ["id", "employee_code", "first_name", "last_name"])
        data["assignee_hierarchy"] = self._employee_hierarchy(self.assignee)
        data["creator_hierarchy"] = self._employee_hierarchy(self.creator)
        data["upload_batch"] = _summary(self.upload_batch, ["id", "file_name", "status"])

        return data

class LeadUploadBatch(TimestampMixin, db.Model):
    __tablename__ = "lead_upload_batches"

    id = db.Column(db.Integer, primary_key=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey("base_users.id"), nullable=False)
    file_name = db.Column(db.String(255), nullable=False)
    total_rows = db.Column(db.Integer, default=0)
    success_count = db.Column(db.Integer, default=0)
    failed_count = db.Column(db.Integer, default=0)
    status = db.Column(db.String(20), default="Processing")  # Processing / Completed / Failed
    error_summary = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    uploader = db.relationship("BaseUser", foreign_keys=[uploaded_by])
    leads = db.relationship("Lead", back_populates="upload_batch")

    def to_dict(self):
        data = super().to_dict()
        data["uploader"] = _summary(self.uploader, ["id", "username", "email"])
        return data


class LeadAssignmentHistory(TimestampMixin, db.Model):
    __tablename__ = "lead_assignment_history"

    id = db.Column(db.Integer, primary_key=True)
    lead_id = db.Column(db.Integer, db.ForeignKey("leads.id"), nullable=False)
    assigned_to = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    assigned_by = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    previous_assignee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    is_active = db.Column(db.Boolean, default=True)

    lead = db.relationship("Lead")
    assignee = db.relationship("Employee", foreign_keys=[assigned_to])
    assigner = db.relationship("Employee", foreign_keys=[assigned_by])
    previous_assignee = db.relationship("Employee", foreign_keys=[previous_assignee_id])

    def to_dict(self):
        data = super().to_dict()
        data["assignee"] = _summary(self.assignee, ["id", "employee_code", "first_name", "last_name"])
        data["assigner"] = _summary(self.assigner, ["id", "employee_code", "first_name", "last_name"])
        data["previous_assignee"] = _summary(
            self.previous_assignee, ["id", "employee_code", "first_name", "last_name"]
        )
        return data


class LeadStatusHistory(TimestampMixin, db.Model):
    __tablename__ = "lead_status_history"

    id = db.Column(db.Integer, primary_key=True)
    lead_id = db.Column(db.Integer, db.ForeignKey("leads.id"), nullable=False)
    old_status = db.Column(db.String(30), nullable=True)
    new_status = db.Column(db.String(30), nullable=False)
    changed_by = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    is_active = db.Column(db.Boolean, default=True)

    lead = db.relationship("Lead")
    changer = db.relationship("Employee", foreign_keys=[changed_by])

    def to_dict(self):
        data = super().to_dict()
        data["changer"] = _summary(self.changer, ["id", "employee_code", "first_name", "last_name"])
        return data

class LeadWeeklySnapshot(TimestampMixin, db.Model):
    """One row per (lead, week_start_date) - captures a lead's
    status/assignment/follow-up state as of that week, so lead
    progress can be reported on a weekly cadence the same way
    DepartmentHeadcount reports employee counts weekly."""

    __tablename__ = "lead_weekly_snapshots"

    id = db.Column(db.Integer, primary_key=True)
    lead_id = db.Column(db.Integer, db.ForeignKey("leads.id"), nullable=False)
    week_start_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(30), nullable=False)
    assigned_to = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    follow_up_count = db.Column(db.Integer, default=0)
    notes = db.Column(db.Text, nullable=True)
    recorded_by = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    is_active = db.Column(db.Boolean, default=True)

    lead = db.relationship("Lead")
    assignee = db.relationship("Employee", foreign_keys=[assigned_to])
    recorder = db.relationship("Employee", foreign_keys=[recorded_by])

    __table_args__ = (
        db.UniqueConstraint("lead_id", "week_start_date", name="uq_lead_weekly_snapshot"),
    )

    def to_dict(self):
        data = super().to_dict()
        data["lead"] = _summary(self.lead, ["id", "lead_name", "status"])
        data["assignee"] = _summary(self.assignee, ["id", "employee_code", "first_name", "last_name"])
        data["recorder"] = _summary(self.recorder, ["id", "employee_code", "first_name", "last_name"])
        return data

    @classmethod
    def generate_for_week(cls, week_start_date):
        """Snapshots every active lead's current state for the given
        week. Upserts - safe to re-run for the same week to refresh
        counts (e.g. follow_up_count) without duplicating rows."""
        from models import FollowUp, Lead

        leads = Lead.query.filter(Lead.is_active == True).all()
        results = []

        for lead in leads:
            follow_up_count = FollowUp.query.join(
                __import__("models").Customer,
                FollowUp.customer_id == __import__("models").Customer.id,
            ).filter(
                __import__("models").Customer.lead_id == lead.id,
                FollowUp.is_active == True,
            ).count()

            snapshot = cls.query.filter_by(lead_id=lead.id, week_start_date=week_start_date).first()
            if not snapshot:
                snapshot = cls(lead_id=lead.id, week_start_date=week_start_date)
                db.session.add(snapshot)

            snapshot.status = lead.status
            snapshot.assigned_to = lead.assigned_to
            snapshot.follow_up_count = follow_up_count
            results.append(snapshot)

        db.session.commit()
        return results


class Customer(TimestampMixin, db.Model):
    __tablename__ = "customers"

    id = db.Column(db.Integer, primary_key=True)
    lead_id = db.Column(db.Integer, db.ForeignKey("leads.id"), nullable=True)
    customer_name = db.Column(db.String(150), nullable=False)
    contact_number = db.Column(db.String(15))
    email = db.Column(db.String(150))
    address = db.Column(db.Text)
    registered_by = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)  # NEW
    is_active = db.Column(db.Boolean, default=True)
    lead = db.relationship("Lead", back_populates="customers")
    registered_by_employee = db.relationship("Employee", foreign_keys=[registered_by])  # NEW

    def to_dict(self):
        data = super().to_dict()
        data["lead"] = _summary(self.lead, ["id", "lead_name", "status"])
        data["registered_by_employee"] = _summary(  # NEW
            self.registered_by_employee, ["id", "employee_code", "first_name", "last_name"]
        )
        return data


class FollowUp(TimestampMixin, db.Model):
    __tablename__ = "follow_ups"

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)
    follow_up_date = db.Column(db.DateTime, nullable=False)
    notes = db.Column(db.Text)
    stage = db.Column(db.String(20), default="Monthly")  # Monthly / 3-Month / 6-Month
    is_completed = db.Column(db.Boolean, default=False)
    next_due_date = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    customer = db.relationship("Customer")

    def to_dict(self):
        data = super().to_dict()
        data["customer"] = _summary(self.customer, ["id", "customer_name"])
        return data


class MembershipPlan(TimestampMixin, db.Model):
    """Manageable CRM membership plan (Silver/Gold/Diamond) with a price
    — admin-editable via /crm/membership-plans. `Meeting.membership_plan`
    stores the plan NAME (not a FK) so historical registrations keep
    their label even if a plan is later renamed/deactivated; the
    incentive engine looks up the current `rate` by name at calculation
    time."""

    __tablename__ = "membership_plans"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(30), nullable=False, unique=True)
    rate = db.Column(db.Numeric(12, 2), nullable=False, default=0)
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        data = super().to_dict()
        if self.rate is not None:
            data["rate"] = float(self.rate)
        return data


class Meeting(TimestampMixin, db.Model):
    """Backs the CRM "Registration" screen. `registered_by` is the CRM
    employee who added the registration (used to count today's/this
    month's registrations for the incentive engine and the CRM employee
    dashboard); `membership_plan` is the Silver/Gold/Diamond plan name
    chosen at registration time (see MembershipPlan for pricing)."""

    __tablename__ = "meetings"

    MEMBERSHIP_PLANS = ("Silver", "Gold", "Diamond")

    @classmethod
    def membership_plan_options(cls):
        """Active MembershipPlan names, falling back to the static tuple
        above only if the table is empty (e.g. before the seed migration
        has run)."""
        names = [
            row.name
            for row in MembershipPlan.query.filter_by(is_active=True).all()
        ]
        return tuple(names) if names else cls.MEMBERSHIP_PLANS

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)
    meeting_date = db.Column(db.DateTime, nullable=False)
    notes = db.Column(db.Text)
    registered_by = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    membership_plan = db.Column(db.String(20), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    customer = db.relationship("Customer")
    registered_by_employee = db.relationship("Employee", foreign_keys=[registered_by])

    def to_dict(self):
        data = super().to_dict()
        data["customer"] = _summary(self.customer, ["id", "customer_name"])
        data["registered_by_employee"] = _summary(
            self.registered_by_employee, ["id", "employee_code", "first_name", "last_name"]
        )
        return data


class IncentiveSlab(TimestampMixin, db.Model):
    """Legacy flat-amount slab used only by EmployeeIncentive.calculate_for_period
    (the older /employee-incentives "Incentive Payouts" screen) — unrelated to
    the newer CRM Incentives engine (incentive_engine.py), which computes 6%
    of a MembershipPlan's rate per qualifying registration instead of a fixed
    slab amount. `period_type` groups slabs into the three ranges an admin
    manages on the Incentive Slabs screen (Weekly 0-10/11+, Monthly 0-40/41+,
    Quarterly 0-120/121+) — kept for display/legacy-payout parity with the
    new engine's period targets, even though only "Monthly" slabs are
    actually consulted by calculate_for_period today."""

    __tablename__ = "incentive_slabs"

    PERIOD_TYPES = ("Weekly", "Monthly", "Quarterly")

    id = db.Column(db.Integer, primary_key=True)
    period_type = db.Column(db.String(20), nullable=False, default="Monthly")
    min_customers = db.Column(db.Integer, nullable=False)
    max_customers = db.Column(db.Integer, nullable=True)  # NULL = no upper bound
    incentive_amount = db.Column(db.Numeric(12, 2), nullable=False)
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return super().to_dict()


class EmployeeTarget(TimestampMixin, db.Model):
    """Registered-customer quota for a CRM employee, set on a Weekly,
    Monthly, or Quarterly cadence. Only the field(s) relevant to
    period_type are populated:
      - Weekly:    week_start_date (year derived from it)
      - Monthly:   month + year
      - Quarterly: quarter (1-4) + year
    Only registrations above this target count as 'additional' for
    incentive calculation."""

    __tablename__ = "employee_targets"

    DEFAULT_TARGET = 0
    PERIOD_TYPES = ("Weekly", "Monthly", "Quarterly")

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    period_type = db.Column(db.String(20), nullable=False, default="Monthly")
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=True)          # Monthly only, 1-12
    quarter = db.Column(db.Integer, nullable=True)         # Quarterly only, 1-4
    week_start_date = db.Column(db.Date, nullable=True)    # Weekly only
    target_customer_count = db.Column(db.Integer, nullable=False, default=0)
    is_active = db.Column(db.Boolean, default=True)

    employee = db.relationship("Employee")

    __table_args__ = (
        db.UniqueConstraint(
            "employee_id", "period_type", "year", "month", "quarter", "week_start_date",
            name="uq_employee_target_period",
        ),
    )

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data

class EmployeeIncentive(TimestampMixin, db.Model):
    __tablename__ = "employee_incentives"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    target_customer_count = db.Column(db.Integer, default=0)   # NEW - snapshot of the limit used
    actual_customer_count = db.Column(db.Integer, default=0)   # NEW - total registered that month
    eligible_customer_count = db.Column(db.Integer, default=0)  # extra beyond target (unchanged name)
    calculated_amount = db.Column(db.Numeric(12, 2), default=0)
    calculation_date = db.Column(db.Date, nullable=True)
    status = db.Column(db.String(20), default="Pending")
    is_active = db.Column(db.Boolean, default=True)

    employee = db.relationship("Employee")

    __table_args__ = (
        db.UniqueConstraint("employee_id", "month", "year", name="uq_employee_incentive_period"),
    )

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        return data

    @classmethod
    def calculate_for_period(cls, month, year):
        """Only considers employees in the CRM department. For each,
        counts customers they registered that month, subtracts their
        target (max limit) to get the extra count, and looks up the
        incentive amount from IncentiveSlab based on that extra count
        (not the raw total). Employees with no extras beyond target
        get no EmployeeIncentive row."""
        from calendar import monthrange
        from models import Customer, Department, Employee

        days_in_month = monthrange(year, month)[1]
        period_start = date(year, month, 1)
        period_end = date(year, month, days_in_month)

        slabs = IncentiveSlab.query.filter(
            IncentiveSlab.is_active == True,
            IncentiveSlab.period_type == "Monthly",
        ).order_by(
            IncentiveSlab.min_customers
        ).all()

        # Scope to CRM department employees only — same "CRM" department
        # name match used on the frontend (CrmEmployeeView).
        crm_employee_ids = {
            row.id
            for row in Employee.query.join(Department).filter(
                db.func.lower(db.func.trim(Department.department_name)) == "crm",
                Employee.is_active == True,
            ).all()
        }

        counts = (
            db.session.query(Customer.registered_by, db.func.count(Customer.id))
            .filter(
                Customer.registered_by.in_(crm_employee_ids) if crm_employee_ids else False,
                Customer.created_at >= period_start,
                Customer.created_at < period_end + timedelta(days=1),
                Customer.is_active == True,
            )
            .group_by(Customer.registered_by)
            .all()
        )

        results = []
        for employee_id, actual_count in counts:
            target_row = EmployeeTarget.query.filter_by(
                employee_id=employee_id, month=month, year=year, is_active=True
            ).first()
            target = target_row.target_customer_count if target_row else EmployeeTarget.DEFAULT_TARGET

            extra_count = max(0, actual_count - target)

            if extra_count <= 0:
                continue  # no extras beyond limit — not incentive-eligible

            amount = 0
            for slab in slabs:
                if extra_count >= slab.min_customers and (
                    slab.max_customers is None or extra_count <= slab.max_customers
                ):
                    amount = float(slab.incentive_amount)
                    break

            if amount <= 0:
                continue  # extras exist but don't reach the lowest slab

            record = EmployeeIncentive.query.filter_by(
                employee_id=employee_id, month=month, year=year
            ).first()

            if not record:
                record = EmployeeIncentive(employee_id=employee_id, month=month, year=year)
                db.session.add(record)

            record.target_customer_count = target
            record.actual_customer_count = actual_count
            record.eligible_customer_count = extra_count
            record.calculated_amount = amount
            record.calculation_date = date.today()
            if record.status == "Pending":
                record.status = "Approved"

            results.append(record)

        db.session.commit()
        return results


# ==================================================================
#  TIER-BASED CRM INCENTIVES
#  IncentiveTier   -> global Bronze / Silver / Gold config
#  WeeklyIncentive -> one row per employee per ISO week
#  MonthlyPayout   -> sum of that month's weekly rows
#  YearlyPayout    -> sum of the 12 monthly payouts
#  Engine: api.v1.crm.incentive_engine
# ==================================================================

class IncentiveTier(TimestampMixin, db.Model):
    __tablename__ = "incentive_tiers"

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(30), nullable=False, unique=True)   # Bronze / Silver / Gold
    # Minimum registrations completed in a week to sit in this tier.
    min_registrations = db.Column(db.Integer, nullable=False, default=0)
    # Paid per incentive-eligible registration while in this tier.
    rate_per_registration = db.Column(db.Numeric(12, 2), nullable=False, default=0)
    sort_order = db.Column(db.Integer, nullable=False, default=0)
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        data = super().to_dict()
        if self.rate_per_registration is not None:
            data["rate_per_registration"] = float(self.rate_per_registration)
        return data

    @classmethod
    def resolve(cls, weekly_count):
        """Highest active tier whose min_registrations <= weekly_count.
        Returns None when the count is below the lowest tier."""
        tiers = (
            cls.query.filter(cls.is_active == True)
            .order_by(cls.min_registrations.desc())
            .all()
        )
        for tier in tiers:
            if weekly_count >= (tier.min_registrations or 0):
                return tier
        return None


class WeeklyIncentive(TimestampMixin, db.Model):
    __tablename__ = "weekly_incentives"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    week_start_date = db.Column(db.Date, nullable=False)   # Monday
    week_end_date = db.Column(db.Date, nullable=False)     # Sunday
    iso_year = db.Column(db.Integer, nullable=False)
    iso_week = db.Column(db.Integer, nullable=False)

    registration_count = db.Column(db.Integer, nullable=False, default=0)
    target_count = db.Column(db.Integer, nullable=False, default=0)
    eligible_count = db.Column(db.Integer, nullable=False, default=0)

    tier_id = db.Column(db.Integer, db.ForeignKey("incentive_tiers.id"), nullable=True)
    tier_name = db.Column(db.String(30), nullable=True)                 # snapshot
    rate_per_registration = db.Column(db.Numeric(12, 2), nullable=False, default=0)  # snapshot
    amount = db.Column(db.Numeric(12, 2), nullable=False, default=0)

    status = db.Column(db.String(20), default="Pending")   # Pending / Approved / Paid
    is_active = db.Column(db.Boolean, default=True)

    employee = db.relationship("Employee")
    tier = db.relationship("IncentiveTier")

    __table_args__ = (
        db.UniqueConstraint("employee_id", "week_start_date", name="uq_weekly_incentive_period"),
    )

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        for k in ("rate_per_registration", "amount"):
            if data.get(k) is not None:
                data[k] = float(data[k])
        return data


class MonthlyPayout(TimestampMixin, db.Model):
    __tablename__ = "monthly_payouts"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)

    week_count = db.Column(db.Integer, nullable=False, default=0)
    registration_count = db.Column(db.Integer, nullable=False, default=0)
    # Snapshot of the Monthly period target used for this row's eligible_count
    # (incentive_engine.PERIOD_TARGETS["Monthly"], or the employee's own
    # EmployeeTarget override) — lets the UI show real progress instead of a
    # hardcoded number.
    target_count = db.Column(db.Integer, nullable=False, default=0)
    eligible_count = db.Column(db.Integer, nullable=False, default=0)
    amount = db.Column(db.Numeric(12, 2), nullable=False, default=0)
    # Incentive is marked payable on the 20th of the month following the
    # incentive period — snapshotted here so the invoice generated from it
    # carries the same due date.
    due_date = db.Column(db.Date, nullable=True)

    status = db.Column(db.String(20), default="Pending")
    is_active = db.Column(db.Boolean, default=True)

    employee = db.relationship("Employee")

    __table_args__ = (
        db.UniqueConstraint("employee_id", "month", "year", name="uq_monthly_payout_period"),
    )

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        if data.get("amount") is not None:
            data["amount"] = float(data["amount"])
        return data


class YearlyPayout(TimestampMixin, db.Model):
    __tablename__ = "yearly_payouts"

    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    year = db.Column(db.Integer, nullable=False)

    month_count = db.Column(db.Integer, nullable=False, default=0)
    registration_count = db.Column(db.Integer, nullable=False, default=0)
    eligible_count = db.Column(db.Integer, nullable=False, default=0)
    amount = db.Column(db.Numeric(12, 2), nullable=False, default=0)

    status = db.Column(db.String(20), default="Pending")
    is_active = db.Column(db.Boolean, default=True)

    employee = db.relationship("Employee")

    __table_args__ = (
        db.UniqueConstraint("employee_id", "year", name="uq_yearly_payout_period"),
    )

    def to_dict(self):
        data = super().to_dict()
        data["employee"] = _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
        if data.get("amount") is not None:
            data["amount"] = float(data["amount"])
        return data


class IncentivePayoutRun(TimestampMixin, db.Model):
    """One row per (month, year) incentive period once its automated
    payout has run — the idempotency guard for
    incentive_engine.auto_process_due_payouts(), so a period is never
    invoiced/paid twice even though the check re-runs on every request
    on/after the 20th."""

    __tablename__ = "incentive_payout_runs"

    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    invoices_created = db.Column(db.Integer, default=0)
    payments_created = db.Column(db.Integer, default=0)
    ran_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint("month", "year", name="uq_incentive_payout_run_period"),
    )


class Quotation(TimestampMixin, db.Model):
    __tablename__ = "quotations"

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=False)
    quotation_number = db.Column(db.String(30), unique=True)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    status = db.Column(db.String(20), default="Draft")
    is_active = db.Column(db.Boolean, default=True)
    customer = db.relationship("Customer")

    def to_dict(self):
        data = super().to_dict()
        data["customer"] = _summary(self.customer, ["id", "customer_name"])
        return data


class Invoice(TimestampMixin, db.Model):
    __tablename__ = "invoices"

    id = db.Column(db.Integer, primary_key=True)
    quotation_id = db.Column(db.Integer, db.ForeignKey("quotations.id"), nullable=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=True)
    invoice_type = db.Column(db.String(20), default="Customer")
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    incentive_id = db.Column(db.Integer, db.ForeignKey("employee_incentives.id"), nullable=True)
    monthly_payout_id = db.Column(db.Integer, db.ForeignKey("monthly_payouts.id"), nullable=True)
    invoice_number = db.Column(db.String(30), unique=True)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    due_date = db.Column(db.Date)
    status = db.Column(db.String(20), default="Unpaid")
    is_active = db.Column(db.Boolean, default=True)
    customer = db.relationship("Customer")
    employee = db.relationship("Employee")
    incentive = db.relationship("EmployeeIncentive")
    payments = db.relationship("Payment", back_populates="invoice", cascade="all, delete-orphan")

    def to_dict(self):
        data = super().to_dict()
        data["customer"] = _summary(self.customer, ["id", "customer_name"]) if self.customer else None
        data["employee"] = (
            _summary(self.employee, ["id", "employee_code", "first_name", "last_name"])
            if self.employee
            else None
        )
        data["incentive"] = (
            _summary(self.incentive, ["id", "month", "year", "eligible_customer_count", "calculated_amount"])
            if self.incentive
            else None
        )
        data["paid_amount"] = float(sum((p.amount or 0) for p in self.payments))
        return data

    @classmethod
    def generate_crm_report(cls, from_date=None, to_date=None):
        query = Lead.query
        rows = []
        for lead in query.order_by(Lead.id).all():
            rows.append([lead.id, lead.lead_name, lead.status, lead.source or "", "Lead", "", ""])

        invoice_query = cls.query
        if from_date and to_date:
            invoice_query = invoice_query.filter(cls.due_date.between(from_date, to_date))
        for invoice in invoice_query.order_by(cls.id).all():
            paid = sum((p.amount or 0) for p in invoice.payments)
            rows.append([
                invoice.id,
                invoice.customer.customer_name if invoice.customer else (
                    f"Employee #{invoice.employee_id}" if invoice.employee_id else ""
                ),
                invoice.status,
                "",
                "Invoice",
                float(invoice.amount or 0),
                float(paid),
            ])
        headers = ["ID", "Name", "Status", "Source", "Record Type", "Amount", "Paid Amount"]
        return Attendance._create_workbook("CRM Report", headers, rows)


class Payment(TimestampMixin, db.Model):
    __tablename__ = "payments"

    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey("invoices.id"), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    payment_date = db.Column(db.Date, nullable=False)
    mode = db.Column(db.String(30))
    # Set when this payment was actually sent through a payment gateway
    # (currently: Razorpay) rather than entered/settled manually — see
    # api/v1/crm/razorpay_gateway.py / incentive_engine.py.
    gateway = db.Column(db.String(30))
    gateway_reference = db.Column(db.String(80))
    is_active = db.Column(db.Boolean, default=True)
    invoice = db.relationship("Invoice", back_populates="payments")

    def to_dict(self):
        data = super().to_dict()
        data["invoice"] = _summary(self.invoice, ["id", "invoice_number", "amount", "status"])
        return data


class SupportTicket(TimestampMixin, db.Model):
    __tablename__ = "support_tickets"

    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey("customers.id"), nullable=True)  # CHANGED: was nullable=False
    lead_id = db.Column(db.Integer, db.ForeignKey("leads.id"), nullable=True)  # NEW
    raised_by = db.Column(db.Integer, db.ForeignKey("base_users.id"), nullable=True)  # NEW
    assigned_to = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)  # NEW
    subject = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default="Open")
    is_active = db.Column(db.Boolean, default=True)
    customer = db.relationship("Customer")
    lead = db.relationship("Lead")  # NEW
    raised_by_user = db.relationship("BaseUser", foreign_keys=[raised_by])  # NEW
    assignee = db.relationship("Employee", foreign_keys=[assigned_to])  # NEW

    def to_dict(self):
        data = super().to_dict()
        data["customer"] = _summary(self.customer, ["id", "customer_name"]) if self.customer else None
        data["lead"] = _summary(self.lead, ["id", "lead_name", "status"]) if self.lead else None  # NEW
        data["raised_by_user"] = _summary(self.raised_by_user, ["id", "username", "email"])  # NEW
        data["assignee"] = _summary(  # NEW
            self.assignee, ["id", "employee_code", "first_name", "last_name"]
        ) if self.assignee else None
        return data


class FeedbackTicket(TimestampMixin, db.Model):
    """HRMS support ticket raised by an authenticated user."""

    __tablename__ = "feedback_tickets"

   
    MAIN_CATEGORIES = (
        "Feature Bug",
        "Internal Bug",
        "Other Bugs/Issues",
    )

    CATEGORIES = MAIN_CATEGORIES

    SUBCATEGORIES = (
        "Login / Password Issue",
        "Account Locked / Access Issue",
        "Employee Profile Update",
        "Employee Master Data Correction",
        "New Employee Creation",
        "Employee Exit / Deactivation",
        "Attendance Issue",
        "Attendance Regularization",
        "Leave Balance Issue",
        "Leave Application Issue",
        "Leave Approval Issue",
        "Holiday / Calendar Issue",
        "Shift / Roster Issue",
        "Work From Home / Remote Work Issue",
        "Overtime Issue",
        "Payroll / Salary Issue",
        "Payslip Issue",
        "Tax / TDS Issue",
        "Reimbursement Issue",
        "Expense Claim Issue",
        "Loan / Advance Issue",
        "Bank Account / Payment Details Update",
        "Benefits / Insurance Issue",
        "Performance Management Issue",
        "Appraisal / Rating Issue",
        "Training / Learning Issue",
        "Recruitment / Hiring Issue",
        "Onboarding Issue",
        "Employee Documents Issue",
        "HR Letter / Certificate Request",
        "Organization / Department Change",
        "Manager / Reporting Structure Change",
        "Transfer / Location Change",
        "Notification / Email Issue",
        "Mobile App Issue",
        "HRMS System Error",
        "Data / Report Issue",
        "Integration Issue",
        "Approval Workflow Issue",
        "Permission / Role Access Request",
        "Feature / Configuration Request",
        "HR Policy / Process Clarification",
        "General HRMS Query",
        "Other / Miscellaneous",
    )

    STATUSES = (
        "Open",
        "In Progress",
        "Resolved",
    )

    id = db.Column(db.Integer, primary_key=True)
    ticket_number = db.Column(db.String(20), unique=True, nullable=True)

    # User account that actually raised the ticket.
    raised_by = db.Column(
        db.Integer,
        db.ForeignKey("base_users.id"),
        nullable=False,
    )

    # Employee record connected to the logged-in user.
    employee_id = db.Column(
        db.Integer,
        db.ForeignKey("employees.id"),
        nullable=True,
    )
    category = db.Column(db.String(120), nullable=False)
    subcategory = db.Column(db.String(120), nullable=True)

    purpose = db.Column(db.String(255), nullable=True)
    description = db.Column(db.Text, nullable=False)
    screenshot_url = db.Column(db.String(500), nullable=True)

    status = db.Column(
        db.String(20),
        nullable=False,
        default="Open",
    )

    # Admin's latest action / resolution note.
    admin_response = db.Column(db.Text, nullable=True)

    resolved_by = db.Column(
        db.Integer,
        db.ForeignKey("base_users.id"),
        nullable=True,
    )

    resolved_at = db.Column(
        db.DateTime,
        nullable=True,
    )

    is_active = db.Column(
        db.Boolean,
        nullable=False,
        default=True,
    )

    raised_by_user = db.relationship(
        "BaseUser",
        foreign_keys=[raised_by],
    )

    resolved_by_user = db.relationship(
        "BaseUser",
        foreign_keys=[resolved_by],
    )

    employee = db.relationship(
        "Employee",
        foreign_keys=[employee_id],
    )

    history = db.relationship(
        "SupportTicketHistory",
        back_populates="ticket",
        cascade="all, delete-orphan",
        order_by="SupportTicketHistory.id.asc()",
    )

    def to_dict(self):
        data = super().to_dict()

        data["raised_by_user"] = _summary(
            self.raised_by_user,
            ["id", "username", "email", "role"],
        )

        data["resolved_by_user"] = (
            _summary(
                self.resolved_by_user,
                ["id", "username", "email"],
            )
            if self.resolved_by_user
            else None
        )

        employee_name = (
            f"{self.employee.first_name or ''} "
            f"{self.employee.last_name or ''}".strip()
            if self.employee
            else None
        )

        data["employee"] = (
            self.employee.employee_code
            if self.employee
            else (
                self.raised_by_user.username
                if self.raised_by_user
                else None
            )
        )

        data["employee_details"] = (
            _summary(
                self.employee,
                [
                    "id",
                    "employee_code",
                    "first_name",
                    "last_name",
                ],
            )
            if self.employee
            else None
        )

        data["employee_id"] = self.employee_id

        data["name"] = employee_name or (
            self.raised_by_user.username
            if self.raised_by_user
            else None
        )

        data["history"] = [
            item.to_dict()
            for item in self.history
        ]

        return data


class SupportTicketHistory(TimestampMixin, db.Model):
    """Audit trail for support-ticket creation and admin updates."""

    __tablename__ = "support_ticket_history"

    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(
        db.Integer,
        db.ForeignKey("feedback_tickets.id"),
        nullable=False,
    )

    action = db.Column(db.String(50), nullable=False)
    performed_by = db.Column(
        db.Integer,
        db.ForeignKey("base_users.id"),
        nullable=True,
    )
    notes = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)

    ticket = db.relationship(
        "FeedbackTicket",
        foreign_keys=[ticket_id],
        back_populates="history",
    )

    performer = db.relationship(
        "BaseUser",
        foreign_keys=[performed_by],
    )

    def to_dict(self):
        data = super().to_dict()
        data["performer"] = _summary(
            self.performer,
            ["id", "username", "email"],
        )
        return data



# ===========================================================================
# Phase 4 — Finance
# ===========================================================================

class Account(TimestampMixin, db.Model):
    __tablename__ = "accounts"

    id = db.Column(db.Integer, primary_key=True)
    account_name = db.Column(db.String(100), nullable=False)
    account_type = db.Column(db.String(30))
    balance = db.Column(db.Numeric(14, 2), default=0)
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return super().to_dict()


class Vendor(TimestampMixin, db.Model):
    __tablename__ = "vendors"

    id = db.Column(db.Integer, primary_key=True)
    vendor_name = db.Column(db.String(150), nullable=False)
    contact_number = db.Column(db.String(15))
    gstin = db.Column(db.String(20))
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return super().to_dict()


class Expense(TimestampMixin, db.Model):
    __tablename__ = "expenses"

    id = db.Column(db.Integer, primary_key=True)
    account_id = db.Column(db.Integer, db.ForeignKey("accounts.id"), nullable=False)
    vendor_id = db.Column(db.Integer, db.ForeignKey("vendors.id"), nullable=True)
    category = db.Column(db.String(50))
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    expense_date = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    account = db.relationship("Account")
    vendor = db.relationship("Vendor")

    def to_dict(self):
        data = super().to_dict()
        data["account"] = _summary(self.account, ["id", "account_name"])
        data["vendor"] = _summary(self.vendor, ["id", "vendor_name"]) if self.vendor else None
        return data

    @classmethod
    def generate_finance_report(cls, from_date=None, to_date=None):
        rows = []
        expense_query = cls.query
        if from_date and to_date:
            expense_query = expense_query.filter(cls.expense_date.between(from_date, to_date))
        for expense in expense_query.order_by(cls.expense_date).all():
            rows.append([
                expense.expense_date.isoformat(),
                "Expense",
                expense.category or "",
                float(expense.amount or 0),
                expense.account.account_name if expense.account else "",
            ])

        income_query = Income.query
        if from_date and to_date:
            income_query = income_query.filter(Income.income_date.between(from_date, to_date))
        for income in income_query.order_by(Income.income_date).all():
            rows.append([
                income.income_date.isoformat(),
                "Income",
                income.source or "",
                float(income.amount or 0),
                income.account.account_name if income.account else "",
            ])

        headers = ["Date", "Type", "Category / Source", "Amount", "Account"]
        return Attendance._create_workbook("Finance Report", headers, rows)


class Income(TimestampMixin, db.Model):
    __tablename__ = "income"

    id = db.Column(db.Integer, primary_key=True)
    account_id = db.Column(db.Integer, db.ForeignKey("accounts.id"), nullable=False)
    source = db.Column(db.String(100))
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    income_date = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    account = db.relationship("Account")

    def to_dict(self):
        data = super().to_dict()
        data["account"] = _summary(self.account, ["id", "account_name"])
        return data

# ==================================================================
#  ADVANCED ATTENDANCE WORKFLOW
#  - AttendanceSetting: one global config row (late cutoff, required
#    hours, permission cap, fixed break durations).
#  - AttendanceEvent: the ordered check-in / check-out timeline for a
#    day. Working / permission / late / overtime figures on the
#    Attendance row are derived from these by
#    api.v1.attendance.attendance_engine.recompute_attendance().
# ==================================================================

class AttendanceSetting(TimestampMixin, db.Model):
    __tablename__ = "attendance_settings"

    id = db.Column(db.Integer, primary_key=True)
    # Any check-in later than this counts as a late login (reason required).
    work_start_time = db.Column(db.Time, nullable=False, default=time(10, 0))
    # Required net working time per day, in hours.
    required_hours_per_day = db.Column(db.Float, nullable=False, default=8.0)
    # Maximum total permission time per day, in minutes.
    max_permission_minutes_per_day = db.Column(db.Integer, nullable=False, default=60)
    # Fixed break durations auto-deducted from gross working time (minutes).
    nap_minutes = db.Column(db.Integer, nullable=False, default=0)
    lunch_minutes = db.Column(db.Integer, nullable=False, default=0)
    tea_minutes = db.Column(db.Integer, nullable=False, default=0)
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        data = super().to_dict()
        if isinstance(self.work_start_time, time):
            data["work_start_time"] = self.work_start_time.strftime("%H:%M")
        return data

    @classmethod
    def get_settings(cls):
        """Return the single global settings row, creating a default one on
        first use so callers never have to null-check."""
        row = cls.query.order_by(cls.id.asc()).first()
        if row is None:
            row = cls()
            db.session.add(row)
            db.session.commit()
        return row


class AttendanceEvent(TimestampMixin, db.Model):
    __tablename__ = "attendance_events"

    id = db.Column(db.Integer, primary_key=True)
    attendance_id = db.Column(
        db.Integer, db.ForeignKey("attendance.id"), nullable=False, index=True
    )
    # "check_in" | "check_out"
    event_type = db.Column(db.String(20), nullable=False)
    event_time = db.Column(db.DateTime, nullable=False)
    # Free-text reason the employee entered for this event.
    reason = db.Column(db.String(255), nullable=True)
    # Why the reason was required:
    #   "late_login"        -> first check-in after work_start_time
    #   "permission"        -> a check-out that starts a permission gap
    #   "permission_return" -> a check-in that ends a permission gap
    #   "overtime"          -> final check-out beyond required hours
    reason_type = db.Column(db.String(30), nullable=True)
    latitude = db.Column(db.Numeric(10, 7))
    longitude = db.Column(db.Numeric(10, 7))
    is_active = db.Column(db.Boolean, default=True)

    attendance = db.relationship("Attendance", back_populates="events")

    def to_dict(self):
        data = super().to_dict()
        if isinstance(self.latitude, (int, float)) or self.latitude is not None:
            data["latitude"] = float(self.latitude) if self.latitude is not None else None
        if self.longitude is not None:
            data["longitude"] = float(self.longitude)
        return data
