"""
Lead upload — two paths, both restricted to admin or a CRM-department
employee whose Designation contains "Marketing" (see
utils.is_crm_marketing_employee):

  POST /lead-uploads/        — a single .xlsx file (multipart field "file"),
                                one Lead per data row.
  POST /lead-uploads/photo   — a single lead photo (multipart field "file",
                                e.g. a phone-gallery picture of handwritten
                                notes), OCR'd via pytesseract to extract a
                                name + contact number into ONE Lead, with
                                the full raw OCR text kept in Lead.notes so
                                a human can verify/correct anything misread.

Both paths share a LeadUploadBatch row (for the existing upload-history
list/deactivate UI) and the same per-lead creation helper so the two
stay in sync.
"""

import re

from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required
from openpyxl import load_workbook

from extensions import db
from models import Lead, LeadUploadBatch
from utils import (
    apply_search_filters,
    fetch_or_404,
    get_current_user,
    is_admin,
    is_crm_marketing_user,
    paginate_query,
    with_token,
)

lead_uploads_bp = Blueprint("lead_uploads_bp", __name__)

ALLOWED_EXTENSIONS = {"xlsx"}
ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "heic", "heif"}
EXPECTED_COLUMNS = ["lead_name", "contact_number", "email", "source", "status"]

# International-ish phone pattern: an optional "+", then 8-15 digits with
# optional spaces/dashes in between — permissive on purpose since OCR text
# is noisy; the extracted match is then stripped down to bare digits (plus
# a leading "+") before being stored.
_PHONE_PATTERN = re.compile(r"\+?\d[\d\s\-]{6,}\d")


def _can_upload_leads(current_user):
    return is_admin(current_user) or is_crm_marketing_user(current_user)


def _allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _allowed_image(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def _build_lead(
    lead_name,
    contact_number=None,
    email=None,
    source="Manual",
    status="New",
    assigned_to=None,
    created_by=None,
    upload_batch_id=None,
    notes=None,
):
    """Single place that shapes a Lead row from raw field values, shared by
    the .xlsx row loop and the photo/OCR path so both stay in sync."""
    return Lead(
        lead_name=str(lead_name).strip()[:150],
        contact_number=(str(contact_number).strip()[:15] if contact_number else None),
        email=(str(email).strip()[:150] if email else None),
        source=(str(source).strip()[:50] if source else "Manual"),
        status=(str(status).strip()[:20] if status else "New"),
        assigned_to=assigned_to,
        created_by=created_by,
        upload_batch_id=upload_batch_id,
        notes=(str(notes).strip() if notes else None),
        is_active=True,
    )


def _extract_lead_fields(raw_text):
    """Best-effort Name / Contact Number extraction from noisy OCR text.

    Contact number: the first phone-shaped run of digits anywhere in the
    text, normalized to digits (and a leading "+") only.

    Name: an explicit "Name:" / "Name -" prefixed line if present, else
    the first non-empty line that isn't itself the matched phone number —
    handwritten notes are usually "Name" on one line and the number on
    another.
    """
    phone_match = _PHONE_PATTERN.search(raw_text or "")
    contact_number = None
    if phone_match:
        contact_number = re.sub(r"[^\d+]", "", phone_match.group(0))

    name = None
    lines = [line.strip() for line in (raw_text or "").splitlines() if line.strip()]

    for line in lines:
        prefixed = re.match(r"(?i)^name\s*[:\-]\s*(.+)$", line)
        if prefixed:
            name = prefixed.group(1).strip()
            break

    if not name:
        for line in lines:
            if phone_match and phone_match.group(0) in line:
                continue
            if re.fullmatch(r"[\d\s+\-]+", line):
                continue
            name = line
            break

    return name, contact_number


@lead_uploads_bp.route("/", methods=["GET"])  # CHANGED: relative
@jwt_required()
@with_token
def list_lead_uploads(token_response):
    query = LeadUploadBatch.query
    query = apply_search_filters(query, request.args, ["file_name", "status"])
    if request.args.get("is_active") is not None:
        query = query.filter(
            LeadUploadBatch.is_active
            == (request.args.get("is_active").lower() in {"true", "1", "yes"})
        )
    return jsonify(
        {
            "message": "Lead upload batches fetched",
            "data": paginate_query(query, request.args),
            "token_response": token_response,
        }
    ), 200


@lead_uploads_bp.route("/<int:batch_id>", methods=["GET"])  # CHANGED: relative
@jwt_required()
@with_token
def get_lead_upload(batch_id, token_response):
    batch, error_response = fetch_or_404(LeadUploadBatch, batch_id)
    if error_response:
        return error_response
    return jsonify(
        {
            "message": "Lead upload batch fetched",
            "data": batch.to_dict(),
            "token_response": token_response,
        }
    ), 200


@lead_uploads_bp.route("/", methods=["POST"])
@jwt_required()
@with_token
def upload_leads(token_response):
    current_user = get_current_user()

    if not _can_upload_leads(current_user):
        return jsonify({"message": "Admin or CRM Marketing privileges required"}), 403

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"message": "No file provided"}), 400

    if not _allowed_file(file.filename):
        return jsonify({"message": "Only .xlsx files are supported"}), 400

    assigned_to_raw = request.form.get("assigned_to")
    assigned_to = None
    if assigned_to_raw:
        try:
            assigned_to = int(assigned_to_raw)
        except (TypeError, ValueError):
            return jsonify({"message": "Invalid assigned_to value"}), 400

    try:
        workbook = load_workbook(file, data_only=True)
    except Exception as exc:
        return jsonify({"message": f"Could not read the uploaded file: {exc}"}), 400

    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    if not rows:
        return jsonify({"message": "The uploaded file has no data rows (only a header, or is empty)"}), 400

    creator_employee_id = None
    try:
        employee = getattr(current_user, "employee", None)
        creator_employee_id = employee.id if employee else None
    except Exception:
        creator_employee_id = None

    batch = LeadUploadBatch(
        uploaded_by=current_user.id,
        file_name=file.filename,
        total_rows=len(rows),
        status="Processing",
    )
    db.session.add(batch)
    db.session.flush()

    success_count = 0
    errors = []

    for index, row in enumerate(rows, start=2):
        try:
            lead_name = row[0] if len(row) > 0 else None
            if not lead_name or not str(lead_name).strip():
                errors.append(f"Row {index}: lead_name is required")
                continue

            lead = _build_lead(
                lead_name=lead_name,
                contact_number=(row[1] if len(row) > 1 else None),
                email=(row[2] if len(row) > 2 else None),
                source=(row[3] if len(row) > 3 and row[3] else "Excel Upload"),
                status=(row[4] if len(row) > 4 and row[4] else "New"),
                assigned_to=assigned_to,
                created_by=creator_employee_id,
                upload_batch_id=batch.id,
            )
            db.session.add(lead)
            success_count += 1
        except Exception as exc:
            errors.append(f"Row {index}: {exc}")
            continue

    batch.success_count = success_count
    batch.failed_count = len(rows) - success_count
    batch.status = "Completed" if batch.failed_count == 0 else "Completed with errors"
    batch.error_summary = "; ".join(errors[:50]) if errors else None

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"message": f"Failed to save uploaded leads: {exc}"}), 500

    return jsonify(
        {
            "message": "Lead upload processed",
            "data": batch.to_dict(),
            "token_response": token_response,
        }
    ), 201


@lead_uploads_bp.route("/photo", methods=["POST"])
@jwt_required()
@with_token
def upload_lead_photo(token_response):
    """OCR a single lead photo (phone-gallery picture of handwritten/typed
    notes) into one Lead. Requires the Tesseract-OCR binary installed on
    the server in addition to the pytesseract/Pillow pip packages —
    see requirements.txt."""
    current_user = get_current_user()

    if not _can_upload_leads(current_user):
        return jsonify({"message": "Admin or CRM Marketing privileges required"}), 403

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"message": "No image provided"}), 400

    if not _allowed_image(file.filename):
        return jsonify({"message": "Only image files (png/jpg/jpeg/webp/heic) are supported"}), 400

    assigned_to_raw = request.form.get("assigned_to")
    assigned_to = None
    if assigned_to_raw:
        try:
            assigned_to = int(assigned_to_raw)
        except (TypeError, ValueError):
            return jsonify({"message": "Invalid assigned_to value"}), 400

    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        return jsonify({
            "message": (
                "OCR support is not installed on this server "
                "(pip install pytesseract Pillow, plus the Tesseract-OCR "
                "binary itself)."
            )
        }), 500

    try:
        image = Image.open(file.stream)
        raw_text = pytesseract.image_to_string(image)
    except Exception as exc:
        return jsonify({"message": f"Could not read text from the image: {exc}"}), 400

    if not raw_text or not raw_text.strip():
        return jsonify({"message": "No readable text was found in the image"}), 400

    name, contact_number = _extract_lead_fields(raw_text)

    if not name:
        return jsonify({
            "message": "Could not identify a lead name in the image — please add it manually",
            "data": {"raw_text": raw_text.strip()},
        }), 422

    creator_employee_id = None
    try:
        employee = getattr(current_user, "employee", None)
        creator_employee_id = employee.id if employee else None
    except Exception:
        creator_employee_id = None

    batch = LeadUploadBatch(
        uploaded_by=current_user.id,
        file_name=file.filename,
        total_rows=1,
        status="Processing",
    )
    db.session.add(batch)
    db.session.flush()

    lead = _build_lead(
        lead_name=name,
        contact_number=contact_number,
        source="Photo Upload",
        status="New",
        assigned_to=assigned_to,
        created_by=creator_employee_id,
        upload_batch_id=batch.id,
        notes=raw_text,
    )
    db.session.add(lead)

    batch.success_count = 1
    batch.failed_count = 0
    batch.status = "Completed"

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"message": f"Failed to save the extracted lead: {exc}"}), 500

    return jsonify({
        "message": "Lead extracted from photo",
        "data": {
            "batch": batch.to_dict(),
            "lead": lead.to_dict(),
            "raw_text": raw_text.strip(),
        },
        "token_response": token_response,
    }), 201


@lead_uploads_bp.route("/<int:batch_id>/deactivate", methods=["DELETE"])  # CHANGED: relative
@jwt_required()
@with_token
def deactivate_lead_upload(batch_id, token_response):
    current_user = get_current_user()

    if not is_admin(current_user):
        return jsonify({"message": "Admin privileges required"}), 403

    batch, error_response = fetch_or_404(LeadUploadBatch, batch_id)
    if error_response:
        return error_response

    if batch.is_active is False:
        return jsonify(
            {
                "message": "Lead upload batch is already inactive",
                "data": batch.to_dict(),
                "token_response": token_response,
            }
        ), 409

    batch.is_active = False
    db.session.commit()
    return jsonify(
        {
            "message": "Lead upload batch deactivated",
            "data": batch.to_dict(),
            "token_response": token_response,
        }
    ), 200