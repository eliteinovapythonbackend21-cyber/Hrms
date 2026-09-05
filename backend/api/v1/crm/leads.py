"""
Lead -> Customer conversion is a single transactional endpoint
(POST /leads/<id>/convert) so a failed second request can never leave an
orphaned Customer without its Lead being marked Converted.
"""

import io
from datetime import timedelta

from flask import jsonify, request, send_file
from flask_jwt_extended import jwt_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from extensions import db
from models import Customer, Lead , Employee , LeadAssignmentHistory, LeadStatusHistory
from utils import (
    fetch_or_404,
    handle_integrity_error,
    is_admin,
    get_current_user,
    parse_date,
    register_crud_blueprint,
    with_token,
    ensure_crm_employee
)


leads_bp = register_crud_blueprint(
    "leads_bp",
    Lead,
    create_fields=[
        "lead_name",
        "contact_number",
        "email",
        "source",
        "status",
        "assigned_to",
        "notes",
        "is_active",
    ],
    search_fields=[
        "lead_name",
        "contact_number",
        "email",
        "source",
        "status",
    ],
    url_prefix_singular="",
    editable=True,
    deletable=False,
    admin_only=False,
)


@leads_bp.route(
    "/<int:lead_id>/deactivate",
    methods=["DELETE"],
)
@jwt_required()
@with_token
def deactivate_lead(
    lead_id,
    token_response,
):
    current_user = get_current_user()

    if not is_admin(current_user):
        return jsonify(
            {
                "message":
                    "Admin privileges required"
            }
        ), 403

    lead, error_response = fetch_or_404(
        Lead,
        lead_id,
    )

    if error_response:
        return error_response

    if lead.is_active is False:
        return jsonify(
            {
                "message":
                    "Lead is already inactive",
                "data":
                    lead.to_dict(),
                "token_response":
                    token_response,
            }
        ), 409

    lead.is_active = False
    db.session.commit()
    return jsonify(
        {
            "message":
                "Lead deactivated",
            "data":
                lead.to_dict(),
            "token_response":
                token_response,
        }
    ), 200


@leads_bp.route(
    "/<int:lead_id>/convert",
    methods=["POST"],
)
@jwt_required()
@with_token
def convert_lead(
    lead_id,
    token_response,
):
    current_user = get_current_user()

    if not is_admin(current_user):
        return jsonify(
            {
                "message":
                    "Admin privileges required"
            }
        ), 403

    lead, error_response = fetch_or_404(
        Lead,
        lead_id,
    )

    if error_response:
        return error_response

    if lead.status == "Converted":
        return jsonify(
            {
                "message":
                    "Lead is already converted"
            }
        ), 409

    if lead.is_active is False:
        return jsonify(
            {
                "message":
                    "Inactive lead cannot be converted"
            }
        ), 400

    data = request.json or {}
    customer = Customer(
        lead_id=lead.id,
        customer_name=(
            data.get("customer_name")
            or lead.lead_name
        ),
        contact_number=(
            data.get("contact_number")
            or lead.contact_number
        ),
        email=(
            data.get("email")
            or lead.email
        ),
        address=data.get(
            "address"
        ),
        is_active=True,
    )

 
    lead.status = "Converted"
    db.session.add(customer)

    try:
        db.session.commit()

    except IntegrityError as exc:
        db.session.rollback()

        return handle_integrity_error(
            exc
        )

    return jsonify(
        {
            "message":
                "Lead converted to customer",
            "data": {
                "lead":
                    lead.to_dict(),
                "customer":
                    customer.to_dict(),
            },
            "token_response":
                token_response,
        }
    ), 201



@leads_bp.route("/<int:lead_id>/assign", methods=["POST"])
@jwt_required()
@with_token
def assign_lead(lead_id, token_response):
    current_user = get_current_user()

    if not is_admin(current_user):
        return jsonify({"message": "Admin privileges required"}), 403

    lead, error_response = fetch_or_404(Lead, lead_id)
    if error_response:
        return error_response

    data = request.json or {}
    new_assignee_id = data.get("assigned_to")

    if not new_assignee_id:
        return jsonify({"message": "assigned_to is required"}), 400

    assignee = Employee.query.get(int(new_assignee_id))
    if not assignee:
        return jsonify({"message": "Employee not found for the provided assigned_to"}), 404

    previous_assignee_id = lead.assigned_to
    assigner_employee = getattr(current_user, "employee", None)

    history = LeadAssignmentHistory(
        lead_id=lead.id,
        assigned_to=assignee.id,
        assigned_by=assigner_employee.id if assigner_employee else None,
        previous_assignee_id=previous_assignee_id,
    )
    lead.assigned_to = assignee.id

    db.session.add(history)

    try:
        db.session.commit()
    except IntegrityError as exc:
        db.session.rollback()
        return handle_integrity_error(exc)

    return jsonify(
        {
            "message": "Lead assigned",
            "data": {"lead": lead.to_dict(), "history": history.to_dict()},
            "token_response": token_response,
        }
    ), 200


@leads_bp.route("/<int:lead_id>/status", methods=["POST"])
@jwt_required()
@with_token
def change_lead_status(lead_id, token_response):
    current_user = get_current_user()

    if not is_admin(current_user):
        return jsonify({"message": "Admin privileges required"}), 403

    lead, error_response = fetch_or_404(Lead, lead_id)
    if error_response:
        return error_response

    data = request.json or {}
    new_status = data.get("status")

    if not new_status:
        return jsonify({"message": "status is required"}), 400

    old_status = lead.status
    changer_employee = getattr(current_user, "employee", None)

    history = LeadStatusHistory(
        lead_id=lead.id,
        old_status=old_status,
        new_status=new_status,
        changed_by=changer_employee.id if changer_employee else None,
    )
    lead.status = new_status

    db.session.add(history)
    db.session.commit()

    return jsonify(
        {
            "message": "Lead status updated",
            "data": {"lead": lead.to_dict(), "history": history.to_dict()},
            "token_response": token_response,
        }
    ), 200


@leads_bp.route("/<int:lead_id>/assignment-history", methods=["GET"])
@jwt_required()
@with_token
def get_lead_assignment_history(lead_id, token_response):
    lead, error_response = fetch_or_404(Lead, lead_id)
    if error_response:
        return error_response

    history = (
        LeadAssignmentHistory.query.filter_by(lead_id=lead.id)
        .order_by(LeadAssignmentHistory.created_at.desc())
        .all()
    )

    return jsonify(
        {
            "message": "Lead assignment history fetched",
            "data": [item.to_dict() for item in history],
            "token_response": token_response,
        }
    ), 200


@leads_bp.route("/<int:lead_id>/status-history", methods=["GET"])
@jwt_required()
@with_token
def get_lead_status_history(lead_id, token_response):
    lead, error_response = fetch_or_404(Lead, lead_id)
    if error_response:
        return error_response

    history = (
        LeadStatusHistory.query.filter_by(lead_id=lead.id)
        .order_by(LeadStatusHistory.created_at.desc())
        .all()
    )

    return jsonify(
        {
            "message": "Lead status history fetched",
            "data": [item.to_dict() for item in history],
            "token_response": token_response,
        }
    ), 200


@leads_bp.route("/report", methods=["GET"])
@jwt_required()
@with_token
def export_lead_generation_report(token_response):
    """Admin-only Lead Generation Report (Excel) — CRM Marketing employees
    can upload leads but never download this, per the access split between
    the two."""
    if not is_admin(get_current_user()):
        return jsonify({"message": "Admin privileges required"}), 403

    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))

    query = Lead.query
    if from_date:
        query = query.filter(Lead.created_at >= from_date)
    if to_date:
        query = query.filter(Lead.created_at < to_date + timedelta(days=1))

    leads = query.order_by(Lead.created_at.desc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lead Generation"

    headers = [
        "Lead ID",
        "Lead Name",
        "Contact Number",
        "Email",
        "Source",
        "Status",
        "Assigned To",
        "Created By",
        "Created At",
    ]
    sheet.append(headers)

    for lead in leads:
        assignee = lead.assignee
        creator = lead.creator
        sheet.append([
            lead.id,
            lead.lead_name or "-",
            lead.contact_number or "-",
            lead.email or "-",
            lead.source or "-",
            lead.status or "-",
            (
                f"{assignee.first_name or ''} {assignee.last_name or ''}".strip()
                if assignee
                else "-"
            ),
            (
                f"{creator.first_name or ''} {creator.last_name or ''}".strip()
                if creator
                else "-"
            ),
            lead.created_at.isoformat() if lead.created_at else "-",
        ])

    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        max_length = max((len(v) for v in values), default=10)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(14, max_length + 2)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="lead_generation_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )