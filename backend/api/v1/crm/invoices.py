"""
Invoice records track billing for a Customer (optionally linked to a
Quotation). Deactivation mirrors the dedicated /deactivate route used
across the rest of the CRM module rather than the generic DELETE
route, which is intentionally blocked (deletable=False).
"""

import io

from flask import jsonify, request, send_file
from flask_jwt_extended import jwt_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from extensions import db
from models import Invoice
from utils import (
    fetch_or_404,
    is_admin,
    get_current_user,
    parse_date,
    register_crud_blueprint,
    with_token,
)


def _generate_invoice_number(item, data):
    if not item.invoice_number:
        next_id = Invoice.get_next_id()
        item.invoice_number = f"INV{next_id:05d}"
    return None


invoices_bp = register_crud_blueprint(
    "invoices_bp",
    Invoice,
    create_fields=[
        "quotation_id",
        "customer_id",
        "invoice_number",
        "amount",
        "due_date",
        "status",
        "is_active",
    ],
    search_fields=[
        "invoice_number",
    ],
    url_prefix_singular="",
    editable=True,
    deletable=False,
    admin_only=False,
    on_create=_generate_invoice_number,
)


@invoices_bp.route(
    "/<int:invoice_id>/deactivate",
    methods=["DELETE"],
)
@jwt_required()
@with_token
def deactivate_invoice(
    invoice_id,
    token_response,
):
    current_user = get_current_user()

    if not is_admin(current_user):
        return jsonify(
            {
                "message":
                    "Admin privileges required"
            }
        ), 403

    invoice, error_response = fetch_or_404(
        Invoice,
        invoice_id,
    )

    if error_response:
        return error_response

    if invoice.is_active is False:
        return jsonify(
            {
                "message":
                    "Invoice is already inactive",
                "data":
                    invoice.to_dict(),
                "token_response":
                    token_response,
            }
        ), 409

    invoice.is_active = False
    db.session.commit()
    return jsonify(
        {
            "message":
                "Invoice deactivated",
            "data":
                invoice.to_dict(),
            "token_response":
                token_response,
        }
    ), 200


@invoices_bp.route("/report", methods=["GET"])
@jwt_required()
@with_token
def export_crm_report(token_response):
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    workbook = Invoice.generate_crm_report(from_date=from_date, to_date=to_date)
    return send_file(
        workbook,
        as_attachment=True,
        download_name="crm_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@invoices_bp.route(
    "/<int:invoice_id>/download",
    methods=["GET"],
)
@jwt_required()
@with_token
def download_invoice(
    invoice_id,
    token_response,
):
    """
    Generates a single-invoice report (.xlsx) for download from the
    UI - the per-card / per-row "Download" action on
    InvoiceListPage.jsx, distinct from /report above (which generates
    an aggregate multi-invoice CRM report across a date range).

    Built directly with openpyxl rather than relying on
    Invoice.generate_crm_report, since that method is designed for a
    multi-row date-range export, not a single-record document.
    """
    invoice, error_response = fetch_or_404(
        Invoice,
        invoice_id,
    )

    if error_response:
        return error_response

    customer = getattr(invoice, "customer", None)
    quotation = getattr(invoice, "quotation", None)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"

    sheet.append(["Field", "Value"])

    rows = [
        (
            "Invoice Number",
            invoice.invoice_number or f"INV{invoice.id:05d}",
        ),
        (
            "Customer",
            getattr(customer, "customer_name", None)
            or f"Customer #{invoice.customer_id}",
        ),
        (
            "Customer Contact",
            getattr(customer, "contact_number", None) or "-",
        ),
        (
            "Customer Email",
            getattr(customer, "email", None) or "-",
        ),
        (
            "Customer Address",
            getattr(customer, "address", None) or "-",
        ),
        (
            "Quotation",
            getattr(quotation, "quotation_number", None)
            or (
                f"Quotation #{invoice.quotation_id}"
                if invoice.quotation_id
                else "No quotation"
            ),
        ),
        (
            "Amount",
            float(invoice.amount) if invoice.amount is not None else 0,
        ),
        (
            "Due Date",
            invoice.due_date.isoformat() if invoice.due_date else "-",
        ),
        (
            "Status",
            invoice.status or "Unpaid",
        ),
        (
            "Active",
            "Yes" if invoice.is_active is not False else "No",
        ),
        (
            "Created At",
            invoice.created_at.isoformat() if getattr(invoice, "created_at", None) else "-",
        ),
    ]

    for label, value in rows:
        sheet.append([label, value])

    for column_cells in sheet.columns:
        values = [
            str(cell.value)
            for cell in column_cells
            if cell.value is not None
        ]
        max_length = max((len(v) for v in values), default=10)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = max(14, max_length + 2)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"invoice_{invoice.invoice_number or invoice.id}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )